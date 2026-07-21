# backend/tests/test_cargas_endpoint.py
"""Endpoints de cargas (Spec §1.6, F-22): POST /cargas (subir extracto) y GET /cargas.

Los caminos de VALIDACIÓN corren en mongomock (fallan antes de tocar la transacción).
El happy path del upload usa transacciones multi-doc → vive en test_carga.py
(@requires_real_mongo, vía el servicio). F-22: solo .xlsx/.xls, .xlsm rechazado,
límite 10 MB.
"""

from decimal import Decimal
from io import BytesIO

import httpx
import openpyxl
import pytest_asyncio
from app.audit.service import configure_audit, reset_audit
from app.auth import passwords, repository
from app.auth.models import User
from app.auth.roles import Role
from app.config import get_settings
from app.domain import DOMAIN_DOCUMENTS
from app.domain.mes_control import MesControl
from app.domain.rubro import Rubro
from app.main import create_app
from beanie import init_beanie
from mongomock_motor import AsyncMongoMockClient

PWD = "clave-larga-1234"


@pytest_asyncio.fixture
async def api(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_ENV", "development")
    monkeypatch.setenv("JWT_SECRET", "x" * 40)
    monkeypatch.setenv("COOKIE_SECURE", "False")
    monkeypatch.setenv("ORIGINALES_DIR", str(tmp_path / "orig"))
    monkeypatch.delenv("RUN_SCHEDULER", raising=False)
    get_settings.cache_clear()

    app = create_app()
    c = AsyncMongoMockClient()
    await init_beanie(database=c["compas_test"], document_models=DOMAIN_DOCUMENTS)
    repository.configure_auth(c, "compas_test")
    configure_audit(c, "compas_test")
    await repository.create_user(
        User(email="fin@roddos.com",
             password_hash=passwords.hash_password(PWD), rol=Role.financiero)
    )
    await Rubro(
        grupo="otros", nombre="Por clasificar", orden=98, es_sistema=True
    ).insert()
    await MesControl(mes="2026-03-01", saldo_inicial_caja=Decimal("0")).insert()

    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac
    repository.reset_auth()
    reset_audit()
    get_settings.cache_clear()


async def _h(ac) -> dict:
    r = await ac.post(
        "/api/v1/auth/login", json={"email": "fin@roddos.com", "password": PWD}
    )
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _xlsx_bbva() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, hdr in enumerate(["FECHA DE OPERACIÓN", "CONCEPTO", "IMPORTE"], start=1):
        ws.cell(row=14, column=i, value=hdr)
    ws.cell(row=15, column=1, value="15-03-2026")
    ws.cell(row=15, column=2, value="COMPRA")
    ws.cell(row=15, column=3, value=-50000)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_xlsm_rechazado(api):
    # F-22: .xlsm (macros) se rechaza SIEMPRE, antes de parsear.
    h = await _h(api)
    r = await api.post(
        "/api/v1/cargas",
        files={"archivo": ("macro.xlsm", b"PK\x03\x04fake", "application/x-xlsm")},
        headers=h,
    )
    assert r.status_code == 422
    assert "xlsm" in r.json()["detail"].lower()


async def test_extension_desconocida_rechazada(api):
    h = await _h(api)
    r = await api.post(
        "/api/v1/cargas",
        files={"archivo": ("datos.csv", b"a,b,c", "text/csv")},
        headers=h,
    )
    assert r.status_code == 422


async def test_limite_10mb(api):
    # F-22: límite de tamaño ANTES de procesar.
    h = await _h(api)
    grande = b"x" * (10 * 1024 * 1024 + 1)
    r = await api.post(
        "/api/v1/cargas",
        files={"archivo": ("ext.xlsx", grande, "application/octet-stream")},
        headers=h,
    )
    assert r.status_code == 413


async def test_sin_auth_401(api):
    r = await api.post(
        "/api/v1/cargas", files={"archivo": ("e.xlsx", b"PK", "application/x")}
    )
    assert r.status_code == 401


async def test_listar_cargas_vacio(api):
    h = await _h(api)
    r = await api.get("/api/v1/cargas", headers=h)
    assert r.status_code == 200
    assert r.json() == {"items": [], "next_cursor": None}
