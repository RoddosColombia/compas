# backend/tests/test_carga.py
"""CargaBancaria (Spec §1.6) + servicio de carga.

MARCADO PARA AUDITORÍA KIMI (flujo crítico: cargas bancarias).

Decisión de contrato registrada: el §1.6 especifica `insertMany ordered=False`
contando duplicados por DuplicateKeyError (idempotente, NO transaccional). La regla 8
lista "finalización de carga" como transacción multi-doc, PERO es incompatible con el
conteo-y-continúa del §1.6 (una transacción abortaría en el 1er duplicado). Se sigue el
§1.6 (data dictionary manda) → pendiente nota/CR y gate Kimi.

Los tests del servicio necesitan Mongo real (índice único parcial + insertMany dedup):
@requires_real_mongo. Los del modelo son puros.
"""

import os
from decimal import Decimal

import openpyxl
import pytest
from app.audit import service as audit_service
from app.audit.events import AuditEvento
from app.domain import DOMAIN_DOCUMENTS
from app.domain.bancos import Banco
from app.domain.carga import CARGAS_COLLECTION, CargaBancaria, EstadoCarga
from app.domain.mes_control import MesControl
from app.domain.rubro import Rubro
from app.domain.transaccion import Transaccion
from beanie import PydanticObjectId, init_beanie
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import ValidationError

# ── Modelo (puro) ─────────────────────────────────────────────────────────


def _carga(**over) -> CargaBancaria:
    base = dict(
        banco=Banco.BBVA,
        archivo_nombre="extracto.xlsx",
        archivo_hash="a" * 64,
        usuario_id=PydanticObjectId(),
    )
    base.update(over)
    return CargaBancaria(**base)


class TestModeloCarga:
    def test_estado_inicial_procesando(self):
        c = _carga()
        assert c.estado is EstadoCarga.PROCESANDO
        assert c.total_filas == 0 and c.nuevas == 0 and c.duplicadas == 0
        assert CargaBancaria.Settings.name == CARGAS_COLLECTION

    def test_banco_manual_no_es_carga(self):
        # Una carga proviene SIEMPRE de un archivo de banco real, nunca 'manual'.
        with pytest.raises(ValidationError):
            _carga(banco=Banco.MANUAL)

    def test_hash_debe_ser_sha256_hex(self):
        with pytest.raises(ValidationError):
            _carga(archivo_hash="no-es-hash")

    def test_rechaza_campo_extra(self):
        with pytest.raises(ValidationError):
            _carga(inventado="x")


# ── Clave de ocurrencia (pura, A-01 + Global66) ──────────────────────────


class TestClaveOcurrencia:
    def _mov(self, **over):
        from datetime import date

        from app.parsers.bank_parsers import MovimientoBancario, TipoMovimiento

        base = dict(
            fecha=date(2026, 7, 14),
            descripcion="Pago PSE",
            monto=Decimal("28373400"),
            tipo=TipoMovimiento.DEBITO,
            banco=Banco.GLOBAL66,
            moneda_original="COP",
            tasa_cambio=Decimal("1"),
            referencia="36666302",
        )
        base.update(over)
        return MovimientoBancario(**base)

    def test_global66_agrupa_por_referencia_no_por_contenido(self):
        # Cargo y su reversa: MISMA referencia, signo/fecha distintos. La clave debe
        # agruparlos (misma operación lógica) → el servicio les da ordinal 1 y 2, así
        # que sus id_banco quedan distintos (ver test_transaccion) y ambos persisten.
        from app.cargas.service import _clave_ocurrencia
        from app.parsers.bank_parsers import TipoMovimiento

        cargo = self._mov(tipo=TipoMovimiento.DEBITO)
        reversa = self._mov(
            tipo=TipoMovimiento.CREDITO, fecha=__import__("datetime").date(2026, 7, 16)
        )
        assert _clave_ocurrencia(cargo) == _clave_ocurrencia(reversa)

    def test_bancolombia_sin_referencia_agrupa_por_contenido(self):
        # Sin referencia nativa (Bancolombia/BBVA): la clave es el contenido; dos
        # movimientos de distinto valor NO comparten clave (huella A-01 intacta).
        from app.cargas.service import _clave_ocurrencia

        a = self._mov(banco=Banco.BANCOLOMBIA, referencia=None)
        b = self._mov(banco=Banco.BANCOLOMBIA, referencia=None, monto=Decimal("99999"))
        assert _clave_ocurrencia(a) != _clave_ocurrencia(b)


# ── Helper de fixture xlsx (BBVA, fechas explícitas) ─────────────────────


def _crear_bbva(path, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(["FECHA DE OPERACIÓN", "CONCEPTO", "IMPORTE"], start=1):
        ws.cell(row=14, column=i, value=h)
    for off, (f, d, v) in enumerate(filas):
        ws.cell(row=15 + off, column=1, value=f)
        ws.cell(row=15 + off, column=2, value=d)
        ws.cell(row=15 + off, column=3, value=v)
    wb.save(str(path))
    wb.close()


# ── Servicio (Mongo real) ────────────────────────────────────────────────


class _StubS3:
    """Cliente S3 stub inyectado en procesar_carga (PR-S3): registra o falla."""

    def __init__(self, fail: bool = False) -> None:
        self.calls: list[dict] = []
        self.fail = fail

    def put_object(self, **kw) -> dict:
        if self.fail:
            raise RuntimeError("s3 caído")
        self.calls.append(kw)
        return {"ETag": "stub"}


@pytest.mark.requires_real_mongo
class TestServicioCarga:
    @pytest.fixture
    async def entorno(self):
        uri = os.environ.get("COMPAS_TEST_MONGO_URI")
        if not uri:
            pytest.skip("COMPAS_TEST_MONGO_URI no configurado")
        client = AsyncIOMotorClient(uri, tz_aware=True)
        dbname = "compas_test_carga"
        await client.drop_database(dbname)
        db = client[dbname]
        await init_beanie(database=db, document_models=DOMAIN_DOCUMENTS)
        audit_service.configure_audit(client, dbname)
        # Semillas mínimas: rubro 'Por clasificar' + MesControl de marzo 2026.
        await Rubro(
            grupo="otros", nombre="Por clasificar", orden=99, es_sistema=True
        ).insert()
        await MesControl(mes="2026-03-01", saldo_inicial_caja=Decimal("0")).insert()
        yield db
        audit_service.reset_audit()
        await client.drop_database(dbname)
        client.close()

    async def _procesar(self, tmp_path, filas, nombre="ext.xlsx"):
        from app.cargas.service import procesar_carga

        p = tmp_path / nombre
        _crear_bbva(p, filas)
        return await procesar_carga(
            banco=Banco.BBVA,
            archivo_path=str(p),
            archivo_nombre=nombre,
            usuario_id=PydanticObjectId(),
            dir_originales=str(tmp_path / "orig"),
        )

    async def test_completada_inserta_transacciones(self, entorno, tmp_path):
        carga = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "COMPRA", -50000),
                ("16-03-2026", "NOMINA", 900000),
            ],
        )
        assert carga.estado is EstadoCarga.COMPLETADA
        assert carga.nuevas == 2
        assert carga.duplicadas == 0
        assert await Transaccion.find(Transaccion.carga_id == carga.id).count() == 2

    async def test_solape_no_duplica(self, entorno, tmp_path):
        # 1ª carga: 1 movimiento. 2ª carga (archivo distinto): el mismo + 1 nuevo.
        await self._procesar(tmp_path, [("15-03-2026", "COMPRA", -50000)], "a.xlsx")
        carga2 = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "COMPRA", -50000),  # solape → duplicado
                ("17-03-2026", "PAGO", -3000),  # nuevo
            ],
            "b.xlsx",
        )
        assert carga2.nuevas == 1
        assert carga2.duplicadas == 1

    async def test_archivo_completado_se_rechaza(self, entorno, tmp_path):
        from app.cargas.service import CargaDuplicadaError, procesar_carga

        p = tmp_path / "dup.xlsx"
        _crear_bbva(p, [("15-03-2026", "COMPRA", -50000)])
        kw = dict(
            banco=Banco.BBVA,
            archivo_path=str(p),
            archivo_nombre="dup.xlsx",
            usuario_id=PydanticObjectId(),
            dir_originales=str(tmp_path / "orig"),
        )
        await procesar_carga(**kw)
        with pytest.raises(CargaDuplicadaError):
            await procesar_carga(**kw)  # mismo hash, ya completada → F-02

    async def test_movimiento_sin_mes_va_a_errores(self, entorno, tmp_path):
        # Abril no tiene MesControl → ese movimiento no se inserta, cuenta como error.
        carga = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "OK MARZO", -1000),
                ("15-04-2026", "SIN MES ABRIL", -2000),
            ],
        )
        assert carga.nuevas == 1
        assert carga.errores == 1
        assert any("2026-04" in e.motivo for e in carga.errores_detalle)

    async def test_emite_evento_carga_completada(self, entorno, tmp_path):
        carga = await self._procesar(tmp_path, [("15-03-2026", "X", -1000)])
        col = entorno["audit_log"]
        doc = await col.find_one({"evento": AuditEvento.carga_completada.value})
        assert doc is not None
        assert doc["entidad_id"] == str(carga.id)

    async def test_identicos_en_un_archivo_no_colapsan(self, entorno, tmp_path):
        # A-01: dos cuotas legítimas idénticas el mismo día → AMBAS entran.
        carga = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "ABONO", -50000),
                ("15-03-2026", "ABONO", -50000),
            ],
        )
        assert carga.nuevas == 2
        assert carga.duplicadas == 0

    async def test_solape_dedup_conserva_identicos(self, entorno, tmp_path):
        # A-01 + dedup: archivo A [X,X]; archivo B (otro hash) [X,X,Z] → solo Z nuevo.
        await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "ABONO", -50000),
                ("15-03-2026", "ABONO", -50000),
            ],
            "a.xlsx",
        )
        carga2 = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "ABONO", -50000),
                ("15-03-2026", "ABONO", -50000),
                ("17-03-2026", "OTRO", -3000),
            ],
            "b.xlsx",
        )
        assert carga2.nuevas == 1
        assert carga2.duplicadas == 2

    async def test_valor_crudo_se_propaga(self, entorno, tmp_path):
        # Regla 7 / Kimi: el texto crudo ambiguo llega al Financiero.
        carga = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "OK", -1000),
                ("15-03-2026", "RARO", "N/A"),
            ],
        )
        assert carga.errores == 1
        assert carga.errores_detalle[0].valor_crudo == "N/A"

    async def test_sin_preservacion_rechaza(self, entorno, tmp_path):
        # M-04 (regla dura): sin S3 ni dir_originales, no se carga.
        from app.cargas.service import OriginalNoPreservableError, procesar_carga

        p = tmp_path / "np.xlsx"
        _crear_bbva(p, [("15-03-2026", "X", -1000)])
        with pytest.raises(OriginalNoPreservableError):
            await procesar_carga(
                banco=Banco.BBVA,
                archivo_path=str(p),
                archivo_nombre="np.xlsx",
                usuario_id=PydanticObjectId(),
            )

    async def test_preserva_original_local(self, entorno, tmp_path):
        from anyio import Path as AsyncPath

        carga = await self._procesar(tmp_path, [("15-03-2026", "X", -1000)])
        assert carga.archivo_s3_key.startswith("local://")
        assert await AsyncPath(carga.archivo_s3_key.removeprefix("local://")).exists()

    # ── PR-S3: preservación del original en S3 (cliente stub inyectado) ──

    async def _procesar_s3(self, tmp_path, filas, client, nombre="ext.xlsx"):
        from app.cargas.service import procesar_carga

        p = tmp_path / nombre
        _crear_bbva(p, filas)
        return await procesar_carga(
            banco=Banco.BBVA,
            archivo_path=str(p),
            archivo_nombre=nombre,
            usuario_id=PydanticObjectId(),
            s3_bucket="compas-archivo",
            s3_client=client,
        )

    async def test_s3_sube_original_y_persiste_key(self, entorno, tmp_path):
        client = _StubS3()
        carga = await self._procesar_s3(
            tmp_path, [("15-03-2026", "COMPRA", -50000)], client
        )
        assert carga.estado is EstadoCarga.COMPLETADA
        assert len(client.calls) == 1  # se subió una vez
        call = client.calls[0]
        assert call["Bucket"] == "compas-archivo"
        assert call["Key"].startswith("originales/")
        # la key persistida es la URI s3://, NO local ni el path temporal
        assert carga.archivo_s3_key == f"s3://compas-archivo/{call['Key']}"

    async def test_s3_fail_closed_no_persiste_nada(self, entorno, tmp_path):
        # put_object falla → la carga aborta ANTES de insertar: ni Carga ni Transaccion.
        from app.cargas.service import procesar_carga

        p = tmp_path / "fc.xlsx"
        _crear_bbva(p, [("15-03-2026", "COMPRA", -50000)])
        client = _StubS3(fail=True)
        with pytest.raises(RuntimeError):
            await procesar_carga(
                banco=Banco.BBVA,
                archivo_path=str(p),
                archivo_nombre="fc.xlsx",
                usuario_id=PydanticObjectId(),
                s3_bucket="compas-archivo",
                s3_client=client,
            )
        assert await CargaBancaria.find_all().count() == 0
        assert await Transaccion.find_all().count() == 0

    async def test_s3_dedup_no_re_sube(self, entorno, tmp_path):
        # F-02 intacto con S3: el mismo archivo (mismo hash) no vuelve a subirse.
        from app.cargas.service import CargaDuplicadaError

        client = _StubS3()
        await self._procesar_s3(
            tmp_path, [("15-03-2026", "COMPRA", -50000)], client, nombre="d.xlsx"
        )
        with pytest.raises(CargaDuplicadaError):
            await self._procesar_s3(
                tmp_path, [("15-03-2026", "COMPRA", -50000)], client, nombre="d.xlsx"
            )
        assert len(client.calls) == 1  # la segunda no subió (dedup antes de S3)

    # ── C3: auto-clasificación al cargar (GO Kimi PLAN-I 9.3, lista §5) ──

    async def _regla(self, patron, rubro, prioridad=10, tipo="egreso"):
        from app.domain.regla_clasificacion import ReglaClasificacion

        r = ReglaClasificacion(
            patron=patron,
            rubro_id=rubro.id,
            tipo_flujo=tipo,
            prioridad=prioridad,
            creada_por="u1",
        )
        await r.insert()
        return r

    async def test_carga_clasifica_con_match(self, entorno, tmp_path):
        # Con match → rubro_id + regla_id escritos (rastro forense §1.5).
        caf = await Rubro(grupo="operacion", nombre="Cafetería", orden=1).insert()
        regla = await self._regla("cafeteria", caf)
        carga = await self._procesar(
            tmp_path, [("15-03-2026", "COMPRA CAFETERÍA LA 14", -50000)]
        )
        assert carga.clasificadas == 1 and carga.por_clasificar == 0
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        assert tx.rubro_id == caf.id
        assert tx.regla_id == regla.id

    async def test_carga_sin_match_cae_a_por_clasificar(self, entorno, tmp_path):
        caf = await Rubro(grupo="operacion", nombre="Cafetería", orden=1).insert()
        await self._regla("cafeteria", caf)
        carga = await self._procesar(
            tmp_path, [("15-03-2026", "GASOLINA TEXACO", -80000)]
        )
        assert carga.clasificadas == 0 and carga.por_clasificar == 1
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        pc = await Rubro.find_one(Rubro.nombre == "Por clasificar")
        assert tx.rubro_id == pc.id
        assert tx.regla_id is None

    async def test_carga_d2_rubro_inactivo_salta_y_reporta(self, entorno, tmp_path):
        # D2 (Kimi): regla activa con rubro inactivo → la fila cae a 'Por
        # clasificar' y la carga reporta reglas_con_rubro_inactivo (fail-loud B-4).
        caf = await Rubro(
            grupo="operacion", nombre="Cafetería", orden=1, activo=False
        ).insert()
        await self._regla("cafeteria", caf)
        carga = await self._procesar(
            tmp_path, [("15-03-2026", "CAFETERIA LA 14", -50000)]
        )
        assert carga.clasificadas == 0 and carga.por_clasificar == 1
        assert carga.reglas_con_rubro_inactivo == ["cafeteria"]
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        pc = await Rubro.find_one(Rubro.nombre == "Por clasificar")
        assert tx.rubro_id == pc.id

    async def test_carga_recorrida_identica(self, entorno, tmp_path):
        # Precedencia determinista: el solape re-cargado (archivo distinto — F-02
        # rechaza el MISMO archivo por hash) queda como duplicado y NO cambia la
        # asignación del original; la fila nueva se clasifica igual.
        caf = await Rubro(grupo="operacion", nombre="Cafetería", orden=1).insert()
        regla = await self._regla("cafeteria", caf)
        await self._procesar(tmp_path, [("15-03-2026", "CAFETERIA X", -1000)], "a.xlsx")
        carga2 = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "CAFETERIA X", -1000),  # solape → duplicado
                ("16-03-2026", "CAFETERIA Y", -2000),  # nueva → misma regla
            ],
            "b.xlsx",
        )
        assert carga2.duplicadas == 1 and carga2.nuevas == 1
        assert carga2.clasificadas == 1  # contadores SOLO sobre las nuevas
        txs = await Transaccion.find_all().to_list()
        assert len(txs) == 2
        assert all(t.rubro_id == caf.id and t.regla_id == regla.id for t in txs)

    async def test_carga_ingreso_clasifica_a_recaudo(self, entorno, tmp_path):
        # D1-ii: partición por tipo — la regla de ingreso ('Abono'→Recaudo) solo
        # ve ingresos; un egreso con texto parecido no se cuela.
        recaudo = await Rubro(
            grupo="otros",
            nombre="Recaudo",
            tipo_flujo="ingreso",
            orden=99,
            es_sistema=True,
        ).insert()
        regla = await self._regla("abono", recaudo, tipo="ingreso")
        carga = await self._procesar(
            tmp_path,
            [
                ("15-03-2026", "ABONO CUOTA SEMANAL", 120000),  # ingreso → Recaudo
                ("16-03-2026", "PAGO ABONO PROVEEDOR", -90000),  # egreso → PC
            ],
        )
        assert carga.clasificadas == 1 and carga.por_clasificar == 1
        ing = await Transaccion.find_one(Transaccion.tipo_flujo == "ingreso")
        egr = await Transaccion.find_one(Transaccion.tipo_flujo == "egreso")
        pc = await Rubro.find_one(Rubro.nombre == "Por clasificar")
        assert ing.rubro_id == recaudo.id and ing.regla_id == regla.id
        assert egr.rubro_id == pc.id and egr.regla_id is None
