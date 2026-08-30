# backend/tests/test_rf_f6_huella_idempotente.py
"""RF-F6 · Fundacional §2 — Cargas idempotentes por huella (antes de Bancolombia sept).

Hoy YA existe (verificado con spec-miner):
  · SHA-256 del archivo en `CargaBancaria.archivo_hash` (calculado en `_hash_archivo`).
  · Storage key = `originales/{hash}{ext}` (S3 Object Lock o dir dev).
  · Dedup por consulta: "hash + estado=completada" → CargaDuplicadaError (409).
  · Dedup de movimientos por índice único parcial `(banco, id_banco)`.

Lo que RF-F6 endurece:
  1. **Candado de BD** — el índice `(archivo_hash, estado)` era NO único. Un race entre
     dos cargas del mismo archivo (A empieza, B empieza antes de que A marque
     COMPLETADA) pasaba las dos por la dedup query. RF-F6 añade el índice único
     parcial `(banco, archivo_hash) WHERE estado=completada` — la BD garantiza que solo
     una puede quedar como COMPLETADA para el mismo (banco, hash).
  2. **Manejo de DuplicateKeyError** — al `save()` que marca COMPLETADA, si el índice
     rechaza, se convierte en `CargaDuplicadaError` con el mismo status que la ruta
     por consulta (idempotente: cliente ve siempre el mismo 409).
  3. **Test end-to-end real-mongo** — dos `procesar_carga` del mismo archivo
     concurrentemente: solo una queda COMPLETADA, la otra recibe CargaDuplicadaError.

Los tests contra el service completo son @requires_real_mongo porque mongomock no
soporta `with_options(read_concern=...)` ni transacciones ni índices únicos parciales.
El test del ÍNDICE en Settings es puro y corre con mongomock — es el canario contra
regresión."""

import asyncio
import os
from decimal import Decimal

import pytest
import pytest_asyncio
from app.audit import service as audit_service
from app.cargas.service import CargaDuplicadaError, procesar_carga
from app.domain import DOMAIN_DOCUMENTS
from app.domain.bancos import Banco
from app.domain.carga import CargaBancaria, EstadoCarga
from app.domain.mes_control import MesControl
from app.domain.rubro import Rubro
from beanie import PydanticObjectId, init_beanie
from motor.motor_asyncio import AsyncIOMotorClient


def test_rff6_indice_unico_parcial_registrado_en_settings():
    """El nuevo candado de BD queda declarado en `CargaBancaria.Settings.indexes`.
    Canario puro (mongomock OK) contra regresión del índice."""
    nombres = {idx.document.get("name") for idx in CargaBancaria.Settings.indexes}
    assert "banco_hash_completada_unico" in nombres, (
        "RF-F6: falta el índice único parcial de idempotencia por huella"
    )


# ─────────────────────── Tests real-mongo ───────────────────────


def _crear_bbva(path, filas: list[tuple[str, str, float]]) -> None:
    """XLSX BBVA mínimo (hoja activa, header en fila 14, según parse_bbva)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(14, 1, "Fecha")
    ws.cell(14, 2, "Concepto")
    ws.cell(14, 3, "Importe")
    for i, (fecha, concepto, importe) in enumerate(filas, start=15):
        ws.cell(i, 1, fecha)
        ws.cell(i, 2, concepto)
        ws.cell(i, 3, importe)
    wb.save(path)


@pytest.mark.requires_real_mongo
class TestIdempotenciaPorHuella:
    @pytest_asyncio.fixture
    async def entorno(self):
        uri = os.environ.get("COMPAS_TEST_MONGO_URI")
        if not uri:
            pytest.skip("COMPAS_TEST_MONGO_URI no configurado")
        client = AsyncIOMotorClient(uri, tz_aware=True)
        dbname = "compas_test_rff6"
        await client.drop_database(dbname)
        db = client[dbname]
        await init_beanie(database=db, document_models=DOMAIN_DOCUMENTS)
        audit_service.configure_audit(client, dbname)
        await Rubro(
            grupo="otros", nombre="Por clasificar", orden=99, es_sistema=True
        ).insert()
        await MesControl(mes="2026-03-01", saldo_inicial_caja=Decimal("0")).insert()
        yield db
        audit_service.reset_audit()
        await client.drop_database(dbname)
        client.close()

    async def _procesar(self, tmp_path, nombre="ext.xlsx"):
        p = tmp_path / nombre
        _crear_bbva(p, [("15-03-2026", "COMPRA", -50000)])
        return await procesar_carga(
            banco=Banco.BBVA,
            archivo_path=str(p),
            archivo_nombre=nombre,
            usuario_id=PydanticObjectId(),
            dir_originales=str(tmp_path / "orig"),
        )

    async def test_mismo_archivo_secuencial_rechazado_por_consulta(
        self, entorno, tmp_path
    ):
        """Camino feliz de F-02: 1ª carga completa, 2ª ve el estado COMPLETADA
        via `_dedup_col.find_one` y se rechaza con CargaDuplicadaError. Este
        camino EXISTÍA — RF-F6 no lo cambia; queda como red de seguridad."""
        c1 = await self._procesar(tmp_path)
        assert c1.estado is EstadoCarga.COMPLETADA
        with pytest.raises(CargaDuplicadaError):
            await self._procesar(tmp_path, nombre="otro.xlsx")

    async def test_dos_cargas_concurrentes_solo_una_completa(self, entorno, tmp_path):
        """El aporte de RF-F6: dos `procesar_carga` del MISMO archivo lanzadas a la
        vez. Sin el índice único parcial, ambas pasarían la dedup por consulta
        (ninguna ve la otra como COMPLETADA aún). CON el índice, solo una gana el
        `save()` que marca COMPLETADA; la otra recibe CargaDuplicadaError."""

        async def correr(nombre):
            try:
                return ("ok", await self._procesar(tmp_path, nombre=nombre))
            except CargaDuplicadaError as e:
                return ("dup", e)

        # Lanzo dos a la vez sobre el MISMO contenido.
        r1, r2 = await asyncio.gather(
            correr("a.xlsx"), correr("b.xlsx"), return_exceptions=False
        )
        # Alguna combinación: (ok, dup) o (dup, ok) — nunca (ok, ok).
        resultados = {r1[0], r2[0]}
        assert "ok" in resultados
        assert "dup" in resultados
        # Y en la BD, exactamente una queda COMPLETADA para ese hash.
        exitosa = r1 if r1[0] == "ok" else r2
        completadas = await CargaBancaria.find(
            CargaBancaria.archivo_hash == exitosa[1].archivo_hash,
            CargaBancaria.estado == EstadoCarga.COMPLETADA,
        ).count()
        assert completadas == 1


if __name__ == "__main__":
    raise SystemExit(pytest.main([__file__, "-q"]))
