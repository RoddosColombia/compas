# backend/tests/test_bank_parsers.py
"""Parsers bancarios (Spec §1.5, portados de SISMO v2 y adaptados a COMPAS).

Cubre las reglas innegociables de CLAUDE.md:
  - Regla 1: montos = Decimal, NUNCA float.
  - Regla 7: los parsers transforman, no interpretan → fila ambigua = error
    acumulado en el resultado, jamás adivinado ni tragado en silencio.
  - Regla 7: Global66 conserva `moneda_original` + `tasa_cambio`.

COMPAS solo opera 3 bancos (enum `Banco`): Bancolombia, BBVA, Global66.
Fixtures sintéticos (openpyxl) — jamás datos reales en el repo.
"""

from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from app.core.time import now_bogota
from app.domain.bancos import Banco
from app.parsers.bank_parsers import (
    ErrorFila,
    MovimientoBancario,
    ResultadoParseo,
    TipoMovimiento,
    detectar_banco,
    parse_bancolombia,
    parse_bbva,
    parse_extracto,
    parse_global66,
)
from pydantic import ValidationError

# ── Helpers de fixtures (estructura real de cada banco) ────────────────


def _crear_bancolombia(path, filas):
    """Bancolombia: hoja 'Extracto', headers fila 15, datos fila 16+.
    `filas` = lista de (fecha_str_d/m, descripcion, valor)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracto"
    for i, h in enumerate(["FECHA", "DESCRIPCIÓN", "VALOR"], start=1):
        ws.cell(row=15, column=i, value=h)
    for off, (f, d, v) in enumerate(filas):
        ws.cell(row=16 + off, column=1, value=f)
        ws.cell(row=16 + off, column=2, value=d)
        ws.cell(row=16 + off, column=3, value=v)
    wb.save(str(path))
    wb.close()


def _crear_bbva(path, filas):
    """BBVA: hoja activa, headers fila 14, datos fila 15+.
    `filas` = lista de (fecha_str_d-m-Y, concepto, importe)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(["FECHA DE OPERACIÓN", "CONCEPTO", "IMPORTE"], start=1):
        ws.cell(row=14, column=i, value=h)
    for off, (f, d, v) in enumerate(filas):
        ws.cell(row=15 + off, column=1, value=f)
        ws.cell(row=15 + off, column=2, value=d)
        ws.cell(row=15 + off, column=3, value=v)
    wb.save(str(path))
    wb.close()


def _crear_global66(path, filas):
    """Global66: hoja 'Movimientos de cuenta COP', headers fila 4, datos fila 5+.
    `filas` = lista de 14 columnas (A..N)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos de cuenta COP"
    ws.cell(row=1, column=1, value="Movimientos de cuenta COP")
    headers = [
        "Tipo transaccion", "Fecha", "Monto debitado", "Monto acreditado",
        "E", "F", "G", "Nombre tercero", "DNI tercero", "J", "K", "L",
        "ID transaccion", "Comentario",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    for off, fila in enumerate(filas):
        for i, val in enumerate(fila, start=1):
            ws.cell(row=5 + off, column=i, value=val)
    wb.save(str(path))
    wb.close()


def _g66_fila(tipo="Debito", fecha="2026-03-15 10:30:00", debito=None,
              credito=None, tercero="TERCERO SA", ref="TXN-001", com="COMENTARIO"):
    return [tipo, fecha, debito, credito, None, None, None, tercero,
            "900123456", None, None, None, ref, com]


# ── Detección de banco ─────────────────────────────────────────────────


class TestDeteccion:
    def test_detecta_bancolombia(self, tmp_path):
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [])
        assert detectar_banco(str(p)) is Banco.BANCOLOMBIA

    def test_detecta_bbva(self, tmp_path):
        p = tmp_path / "v.xlsx"
        _crear_bbva(p, [])
        assert detectar_banco(str(p)) is Banco.BBVA

    def test_detecta_global66(self, tmp_path):
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [])
        assert detectar_banco(str(p)) is Banco.GLOBAL66

    def test_formato_no_soportado_lanza_error(self, tmp_path):
        # COMPAS no soporta PDF/Nequi/Davivienda: debe fallar explícito, no adivinar.
        p = tmp_path / "x.pdf"
        p.write_bytes(b"%PDF-1.4 fake")
        with pytest.raises(ValueError):
            detectar_banco(str(p))


# ── Regla 1: Decimal, nunca float ────────────────────────────────────────


class TestReglaDecimal:
    def test_monto_parseado_es_decimal(self, tmp_path):
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [("15/03", "PAGO", -50000)])
        res = parse_bancolombia(str(p))
        m = res.movimientos[0]
        assert isinstance(m.monto, Decimal)
        assert not isinstance(m.monto, float)

    def test_modelo_rechaza_monto_float(self):
        # El tipo Money debe rechazar float en la construcción del modelo (regla 1).
        with pytest.raises(ValidationError):
            MovimientoBancario(
                fecha=date(2026, 3, 15),
                descripcion="X",
                monto=50000.0,  # float → prohibido
                tipo=TipoMovimiento.DEBITO,
                banco=Banco.BANCOLOMBIA,
            )


# ── Bancolombia ──────────────────────────────────────────────────────────


class TestBancolombia:
    def test_debito_y_credito(self, tmp_path):
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [
            ("15/03", "COMPRA", -50000),
            ("16/03", "ABONO", 120000),
        ])
        res = parse_bancolombia(str(p))
        assert isinstance(res, ResultadoParseo)
        assert res.errores == []
        anio = now_bogota().year
        deb, cred = res.movimientos
        assert deb.tipo is TipoMovimiento.DEBITO
        assert deb.monto == Decimal("50000")
        assert deb.fecha == date(anio, 3, 15)
        assert deb.banco is Banco.BANCOLOMBIA
        assert cred.tipo is TipoMovimiento.CREDITO
        assert cred.monto == Decimal("120000")

    def test_fila_totalmente_vacia_se_omite(self, tmp_path):
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [
            (None, None, None),
            ("17/03", "PAGO", -3000),
        ])
        res = parse_bancolombia(str(p))
        assert len(res.movimientos) == 1
        assert res.errores == []

    def test_monto_invalido_va_a_errores(self, tmp_path):
        # Regla 7: valor ambiguo NO se adivina como 0 ni se traga → error de fila.
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [("18/03", "RARO", "N/A")])
        res = parse_bancolombia(str(p))
        assert res.movimientos == []
        assert len(res.errores) == 1
        assert isinstance(res.errores[0], ErrorFila)
        assert res.errores[0].fila == 16

    def test_fecha_invalida_va_a_errores(self, tmp_path):
        p = tmp_path / "b.xlsx"
        _crear_bancolombia(p, [("nn/nn", "MALA FECHA", -1000)])
        res = parse_bancolombia(str(p))
        assert res.movimientos == []
        assert len(res.errores) == 1


# ── BBVA ─────────────────────────────────────────────────────────────────


class TestBBVA:
    def test_parse_basico(self, tmp_path):
        p = tmp_path / "v.xlsx"
        _crear_bbva(p, [
            ("15-03-2026", "RETIRO", -75000),
            ("16-03-2026", "NOMINA", 900000),
        ])
        res = parse_bbva(str(p))
        assert res.errores == []
        deb, cred = res.movimientos
        assert deb.tipo is TipoMovimiento.DEBITO
        assert deb.monto == Decimal("75000")
        assert deb.fecha == date(2026, 3, 15)
        assert deb.banco is Banco.BBVA
        assert cred.tipo is TipoMovimiento.CREDITO


# ── Global66 (con moneda_original + tasa_cambio, regla 7) ─────────────────


class TestGlobal66:
    def test_egreso_columna_debito(self, tmp_path):
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [_g66_fila(
            tipo="Debito", debito=150000.0, ref="TXN-001", com="ALMUERZO",
            tercero="RESTAURANTE XYZ")])
        res = parse_global66(str(p))
        m = res.movimientos[0]
        assert m.tipo is TipoMovimiento.DEBITO
        assert m.monto == Decimal("150000")
        assert m.fecha == date(2026, 3, 15)
        assert m.banco is Banco.GLOBAL66
        assert m.referencia == "TXN-001"
        assert "ALMUERZO" in m.descripcion
        assert "RESTAURANTE XYZ" in m.descripcion

    def test_ingreso_columna_credito(self, tmp_path):
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [_g66_fila(
            tipo="Abono", debito=None, credito=2500000.0, ref="TXN-002")])
        res = parse_global66(str(p))
        m = res.movimientos[0]
        assert m.tipo is TipoMovimiento.CREDITO
        assert m.monto == Decimal("2500000")
        assert m.referencia == "TXN-002"

    def test_conserva_moneda_y_tasa(self, tmp_path):
        # Regla 7: hoja COP -> moneda_original=COP, tasa_cambio=1 (hecho, no adivinado).
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [_g66_fila(debito=5000.0)])
        m = parse_global66(str(p)).movimientos[0]
        assert m.moneda_original == "COP"
        assert m.tasa_cambio == Decimal("1")
        assert isinstance(m.tasa_cambio, Decimal)

    def test_omite_filas_sin_valor(self, tmp_path):
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [
            _g66_fila(tipo="GMF", debito=None, credito=None, ref="TXN-VACIO"),
            _g66_fila(tipo="Debito", debito=5000.0, ref="TXN-OK"),
        ])
        res = parse_global66(str(p))
        assert len(res.movimientos) == 1
        assert res.movimientos[0].referencia == "TXN-OK"


# ── Dispatcher ───────────────────────────────────────────────────────────


class TestParseExtracto:
    def test_autodetecta_y_rutea(self, tmp_path):
        p = tmp_path / "g.xlsx"
        _crear_global66(p, [_g66_fila(debito=1000.0)])
        res = parse_extracto(str(p))
        assert res.banco is Banco.GLOBAL66
        assert len(res.movimientos) == 1
