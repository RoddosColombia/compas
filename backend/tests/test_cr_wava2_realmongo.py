# backend/tests/test_cr_wava2_realmongo.py
"""CR-WAVA-2 (real-mongo): el hook de clasificación Wava dentro del flujo de carga
(transacción multi-doc) y de `aplicar_pendientes`. La carga usa `with_transaction` →
requiere Mongo real (replica set). Corre en el job backend-real-mongo de CI.

Semántica verificada (idéntica a la trampa de CR-WAVA, pero por vía AUTOMÁTICA):
un depósito Wava con remanente vivo → rubro tránsito → excluido de `ingreso_real`
(no infla recaudo); agotado el remanente → recaudo normal y SÍ cuenta.
"""

import os
from decimal import Decimal

import openpyxl
import pytest
from app.audit import service as audit_service
from app.cierre.transito import RUBRO_TRANSITO, transito_remanente
from app.domain import DOMAIN_DOCUMENTS
from app.domain.bancos import Banco
from app.domain.mes_control import EstadoMes, MesControl
from app.domain.rubro import Rubro, RubroGrupo, TipoFlujo
from app.domain.transaccion import Transaccion
from app.metas_ingreso.service import ingreso_real
from beanie import PydanticObjectId, init_beanie
from motor.motor_asyncio import AsyncIOMotorClient

RUBRO_PC = "Por clasificar"


def _crear_bbva(path, filas):
    """Extracto BBVA mínimo: IMPORTE positivo = crédito = INGRESO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(["FECHA DE OPERACIÓN", "CONCEPTO", "IMPORTE"], start=1):
        ws.cell(row=14, column=i, value=h)
    for off, (f, d, v) in enumerate(filas):
        ws.cell(row=15 + off, column=1, value=f)
        ws.cell(row=15 + off, column=2, value=d)
        ws.cell(row=15 + off, column=3, value=v)
    wb.save(str(path))
    wb.close()


@pytest.mark.requires_real_mongo
class TestHookWavaCarga:
    @pytest.fixture
    async def entorno(self):
        uri = os.environ.get("COMPAS_TEST_MONGO_URI")
        if not uri:
            pytest.skip("COMPAS_TEST_MONGO_URI no configurado")
        client = AsyncIOMotorClient(uri, tz_aware=True)
        dbname = "compas_test_wava2"
        await client.drop_database(dbname)
        db = client[dbname]
        await init_beanie(database=db, document_models=DOMAIN_DOCUMENTS)
        audit_service.configure_audit(client, dbname)
        # Semillas: rubros de sistema + agosto ABIERTO (destino de la carga).
        await Rubro(
            grupo=RubroGrupo.OTROS, nombre=RUBRO_PC, orden=99, es_sistema=True
        ).insert()
        yield db
        audit_service.reset_audit()
        await client.drop_database(dbname)
        client.close()

    async def _rubro_transito(self) -> Rubro:
        return await Rubro(
            grupo=RubroGrupo.OTROS,
            nombre=RUBRO_TRANSITO,
            tipo_flujo=TipoFlujo.INGRESO,
            orden=98,
            es_sistema=True,
        ).insert()

    async def _julio_cerrado(self, transito: str) -> None:
        await MesControl(
            mes="2026-07-01",
            saldo_inicial_caja=Decimal("0"),
            estado=EstadoMes.CERRADO,
            transito_wava=Decimal(transito),
        ).insert()

    async def _agosto_abierto(self) -> MesControl:
        return await MesControl(
            mes="2026-08-01", saldo_inicial_caja=Decimal("0")
        ).insert()

    async def _procesar(self, tmp_path, filas, nombre="ext.xlsx"):
        from app.cargas.service import procesar_carga

        p = tmp_path / nombre
        _crear_bbva(p, filas)
        return await procesar_carga(
            banco=Banco.BBVA,
            archivo_path=str(p),
            archivo_nombre=nombre,
            usuario_id=PydanticObjectId(),
            dir_originales=str(tmp_path / "orig"),
        )

    # ── 1. remanente>0 → tránsito (recaudo inmóvil, total invariante) ──
    async def test_deposito_wava_con_remanente_va_a_transito(self, entorno, tmp_path):
        rt = await self._rubro_transito()
        await self._julio_cerrado("100000")
        await self._agosto_abierto()
        carga = await self._procesar(
            tmp_path, [("10-08-2026", "Recibido de WAVA Technologie", 60000)]
        )
        assert carga.nuevas == 1
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        assert tx.rubro_id == rt.id  # sello de tránsito
        assert tx.regla_id is None  # sistema, no regla
        # No infla recaudo: excluido de ingreso_real (es el único ingreso → 0).
        assert await ingreso_real("2026-08") == Decimal("0")
        # Remanente baja 100k → 40k (la llegada descuenta).
        assert await transito_remanente("2026-08-01") == Decimal("40000")

    # ── 2. remanente agotado → recaudo normal y SÍ cuenta en ingreso_real ──
    async def test_deposito_wava_sin_remanente_es_recaudo(self, entorno, tmp_path):
        rt = await self._rubro_transito()
        await self._julio_cerrado("100000")
        ago = await self._agosto_abierto()
        # Llegada previa que AGOTA el remanente (100k ya aterrizado).
        await Transaccion(
            fecha="2026-08-05",
            descripcion="recibido de wava",
            valor=Decimal("100000"),
            tipo_flujo=TipoFlujo.INGRESO,
            rubro_id=rt.id,
            mes_id=ago.id,
            banco=Banco.GLOBAL66,
            id_banco="WAVA-PREV|1",
        ).insert()
        assert await transito_remanente("2026-08-01") == Decimal("0")
        carga = await self._procesar(
            tmp_path, [("10-08-2026", "Recibido de WAVA Technologie", 70000)]
        )
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        pc = await Rubro.find_one(Rubro.nombre == RUBRO_PC)
        assert tx.rubro_id == pc.id  # sin regla → Por clasificar (no tránsito)
        # SÍ cuenta como recaudo (Por clasificar no es neutro).
        assert await ingreso_real("2026-08") == Decimal("70000")

    # ── 3. sin declaración previa → recaudo siempre ──
    async def test_sin_declaracion_wava_es_recaudo(self, entorno, tmp_path):
        await self._rubro_transito()
        await self._agosto_abierto()  # julio NO cerrado con tránsito
        carga = await self._procesar(
            tmp_path, [("10-08-2026", "Recibido de WAVA Technologie", 50000)]
        )
        tx = await Transaccion.find_one(Transaccion.carga_id == carga.id)
        pc = await Rubro.find_one(Rubro.nombre == RUBRO_PC)
        assert tx.rubro_id == pc.id
        assert await ingreso_real("2026-08") == Decimal("50000")

    # ── 4. no-Wava intacto + coexiste con un Wava en la misma carga ──
    async def test_no_wava_intacto_y_coexiste(self, entorno, tmp_path):
        rt = await self._rubro_transito()
        await self._julio_cerrado("100000")
        await self._agosto_abierto()
        carga = await self._procesar(
            tmp_path,
            [
                ("10-08-2026", "Recibido de WAVA Technologie", 60000),  # → tránsito
                ("11-08-2026", "Recibido de Éxito", 25000),  # → Por clasificar
            ],
        )
        pc = await Rubro.find_one(Rubro.nombre == RUBRO_PC)
        txs = {
            t.descripcion: t
            async for t in Transaccion.find(Transaccion.carga_id == carga.id)
        }
        assert txs["Recibido de WAVA Technologie"].rubro_id == rt.id
        assert txs["Recibido de Éxito"].rubro_id == pc.id
        # Solo el no-Wava cuenta como recaudo.
        assert await ingreso_real("2026-08") == Decimal("25000")

    # ── 5. batch: descuento dentro de un mismo archivo (el 3º cae a recaudo) ──
    async def test_batch_descuenta_y_el_tercero_cae_a_recaudo(self, entorno, tmp_path):
        rt = await self._rubro_transito()
        await self._julio_cerrado("80000")  # remanente 80k
        await self._agosto_abierto()
        carga = await self._procesar(
            tmp_path,
            [
                ("10-08-2026", "Recibido de WAVA Technologie", 60000),  # 80k>0 transito
                ("11-08-2026", "Recibido de WAVA Technologie", 30000),  # 20k>0 transito
                ("12-08-2026", "Recibido de WAVA Technologie", 20000),  # 0 no>0 recaudo
            ],
        )
        pc = await Rubro.find_one(Rubro.nombre == RUBRO_PC)
        por_valor = {
            t.valor: t async for t in Transaccion.find(Transaccion.carga_id == carga.id)
        }
        assert por_valor[Decimal("60000")].rubro_id == rt.id
        assert por_valor[Decimal("30000")].rubro_id == rt.id
        assert por_valor[Decimal("20000")].rubro_id == pc.id  # descuento en batch
        # Solo el 3º (recaudo) cuenta.
        assert await ingreso_real("2026-08") == Decimal("20000")


@pytest.mark.requires_real_mongo
class TestHookWavaAplicarPendientes:
    @pytest.fixture
    async def entorno(self):
        uri = os.environ.get("COMPAS_TEST_MONGO_URI")
        if not uri:
            pytest.skip("COMPAS_TEST_MONGO_URI no configurado")
        client = AsyncIOMotorClient(uri, tz_aware=True)
        dbname = "compas_test_wava2_ap"
        await client.drop_database(dbname)
        db = client[dbname]
        await init_beanie(database=db, document_models=DOMAIN_DOCUMENTS)
        audit_service.configure_audit(client, dbname)
        await Rubro(
            grupo=RubroGrupo.OTROS, nombre=RUBRO_PC, orden=99, es_sistema=True
        ).insert()
        yield db
        audit_service.reset_audit()
        await client.drop_database(dbname)
        client.close()

    async def test_aplicar_pendientes_reclasifica_wava_a_transito(
        self, entorno, tmp_path
    ):
        from app.reglas.service import aplicar_pendientes

        rt = await Rubro(
            grupo=RubroGrupo.OTROS,
            nombre=RUBRO_TRANSITO,
            tipo_flujo=TipoFlujo.INGRESO,
            orden=98,
            es_sistema=True,
        ).insert()
        pc = await Rubro.find_one(Rubro.nombre == RUBRO_PC)
        await MesControl(
            mes="2026-07-01",
            saldo_inicial_caja=Decimal("0"),
            estado=EstadoMes.CERRADO,
            transito_wava=Decimal("100000"),
        ).insert()
        ago = await MesControl(
            mes="2026-08-01", saldo_inicial_caja=Decimal("0")
        ).insert()
        # Depósito Wava atrapado en 'Por clasificar' (sin regla que lo tomara).
        tx = await Transaccion(
            fecha="2026-08-10",
            descripcion="Recibido de WAVA Technologie",
            valor=Decimal("60000"),
            tipo_flujo=TipoFlujo.INGRESO,
            rubro_id=pc.id,
            mes_id=ago.id,
            banco=Banco.GLOBAL66,
            id_banco="WAVA-AP|1",
        ).insert()

        usuario = str(PydanticObjectId())
        res = await aplicar_pendientes(usuario_id=usuario)

        assert res["clasificadas"] == 1
        releido = await Transaccion.get(tx.id)
        assert releido.rubro_id == rt.id  # reclasificado a tránsito
        assert releido.regla_id is None  # sello de sistema
        assert releido.clasificada_por == usuario  # rastro forense B-2
        assert await ingreso_real("2026-08") == Decimal("0")  # excluido
