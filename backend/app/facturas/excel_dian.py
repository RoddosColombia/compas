# backend/app/facturas/excel_dian.py
"""C2' (acta FABS 2026-08-10) — import masivo del Excel de "documentos recibidos"
del portal DIAN: las facturas emitidas A NOMBRE DE RODDOS (gasto con IVA
potencialmente deducible), una fila por documento.

Regla 7 (parsers transforman, NUNCA interpretan), aplicada en tres capas:
  1. CONTRATO DE COLUMNAS explícito (`COLUMNAS`): si los encabezados del archivo
     real no cuadran, el error LISTA esperado vs encontrado — ese mensaje es el
     punto de calibración con el export real, no un fallo silencioso.
  2. Fila ilegible (fecha/monto ambiguos) → error de ESA fila con motivo; las
     demás siguen (resultado parcial, mismo criterio de la ingesta PDF).
  3. Nada se adivina: tipo de documento no soportado (NC/ND) o una fila EMITIDA
     por RODDOS se rechazan con motivo, jamás se ingresan "corregidas".

Persistencia y auditoría: la MISMA ruta que la ingesta PDF (dedup por CUFE +
DuplicateKeyError + `factura.creada` fail-closed saga O1) vía
`ingesta.persistir_factura_ingesta` — un solo camino de escritura.
Deducibilidad: default False/sin decidir (el operador decide, §2 del spec E2);
Auteco por NIT de Configuracion → True/decidida/origen auteco (CEO 2026-07-31).
"""

import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.domain.factura import OrigenFactura, TipoFactura

MAX_FILAS_BUSQUEDA_ENCABEZADO = 15

# ── Contrato de columnas (EL punto de calibración con el export real) ─────────
# clave interna → alias aceptados (normalizados: minúsculas, sin tildes).
# `cufe` matchea por CONTAINS (el portal usa "CUFE/CUDE" o el nombre largo);
# el resto por igualdad exacta contra alguno de sus alias.
COLUMNAS: dict[str, tuple[str, ...]] = {
    "tipo_documento": ("tipo de documento", "tipo documento"),
    "folio": ("folio", "numero de documento", "número de documento"),
    "cufe": ("cufe",),  # contains
    "fecha": ("fecha emision", "fecha de emision"),
    "nit_emisor": ("nit emisor", "nit del emisor"),
    "nombre_emisor": ("nombre emisor", "razon social emisor", "nombre del emisor"),
    "iva": ("iva",),
    "total": ("total", "total factura"),
}
OPCIONALES: dict[str, tuple[str, ...]] = {
    "prefijo": ("prefijo",),
    "nit_receptor": ("nit receptor",),
    "inc": ("inc",),
    "bolsas": ("inc bolsas",),
    "rete_iva": ("rete iva",),
    "rete_fuente": ("rete renta",),
    "rete_ica": ("rete ica",),
}

# Calibrado contra el export REAL del portal (FACTURACION DIAN ENE1-AGO10 2026):
# el tipo viene como "Factura electrónica" (a secas) o "Factura electrónica de
# contingencia"; las notas crédito como "Nota de crédito electrónica". Los
# documentos equivalentes (POS, transporte) NO soportan IVA descontable sin
# factura → se rechazan con motivo.
_TIPO_FACTURA_PREFIJO = "factura electronica"
_MONTO_ES_CO = re.compile(r"^\d{1,3}(\.\d{3})*(,\d+)?$")
_MONTO_PLANO = re.compile(r"^\d+(\.\d+)?$")


class EncabezadosNoReconocidos(Exception):
    """Los encabezados del archivo no cuadran con el contrato (regla 7)."""


class FilaIlegible(Exception):
    """Una fila puntual no se pudo transformar sin interpretar."""


def _norm(v: object) -> str:
    s = str(v or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _mapear_encabezados(celdas: list[object]) -> dict[str, int] | None:
    """Fila de celdas → {clave: índice} si TODAS las requeridas resuelven."""
    normalizadas = [_norm(c) for c in celdas]
    mapa: dict[str, int] = {}
    for clave, alias in {**COLUMNAS, **OPCIONALES}.items():
        for i, h in enumerate(normalizadas):
            if not h:
                continue
            if clave == "cufe":
                if "cufe" in h:
                    mapa[clave] = i
                    break
            elif h in alias:
                mapa[clave] = i
                break
    if all(k in mapa for k in COLUMNAS):
        return mapa
    return None


def _fecha_iso(v: object, fila: int) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    raise FilaIlegible(
        f"fila {fila}: fecha de emisión ilegible ('{s}'); se esperaba una fecha "
        "de Excel o 'YYYY-MM-DD' / 'DD/MM/YYYY'"
    )


def _monto(v: object, campo: str, fila: int) -> Decimal:
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v or "").strip()
    try:
        if _MONTO_ES_CO.match(s):
            return Decimal(s.replace(".", "").replace(",", "."))
        if _MONTO_PLANO.match(s):
            return Decimal(s)
    except InvalidOperation:
        pass
    raise FilaIlegible(
        f"fila {fila}: {campo} ilegible ('{s}'); se esperaba un número o "
        "formato es-CO ('1.452,94')"
    )


def _monto_opcional(v: object, campo: str, fila: int) -> Decimal:
    """Columna opcional: celda vacía/ausente → 0 (el export trae '0' explícito;
    ausencia de la COLUMNA no es dato → 0 sin inventar nada más)."""
    if v in (None, ""):
        return Decimal("0.00")
    return _monto(v, campo, fila)


def parsear_excel(contenido: bytes) -> list[dict]:
    """bytes del .xlsx → filas crudas [{fila, tipo_documento, cufe, numero, ...}].

    Levanta `EncabezadosNoReconocidos` (archivo entero, con esperado vs encontrado)
    o marca por fila con {'error': motivo} — el lote sigue (regla 7 + parcial)."""
    wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active

    mapa: dict[str, int] | None = None
    fila_encabezado = 0
    mejores: list[str] = []
    for i, row in enumerate(ws.iter_rows(max_row=MAX_FILAS_BUSQUEDA_ENCABEZADO), 1):
        celdas = [c.value for c in row]
        mapa = _mapear_encabezados(celdas)
        if mapa is not None:
            fila_encabezado = i
            break
        con_texto = [str(c) for c in celdas if c not in (None, "")]
        if len(con_texto) > len(mejores):
            mejores = con_texto
    if mapa is None:
        esperadas = ", ".join(sorted(a[0] for a in COLUMNAS.values()))
        encontradas = ", ".join(mejores) or "(ninguna)"
        raise EncabezadosNoReconocidos(
            "encabezados no reconocidos en el Excel: se esperaba una fila con "
            f"[{esperadas}] (más 'cufe'); lo más parecido fue [{encontradas}]. "
            "Verifica que sea el export de 'Documentos recibidos' del portal DIAN."
        )

    filas: list[dict] = []
    for n, row in enumerate(
        ws.iter_rows(min_row=fila_encabezado + 1), fila_encabezado + 1
    ):
        celdas = [c.value for c in row]
        if all(c in (None, "") for c in celdas):
            continue

        def celda(clave: str, _celdas: list[object] = celdas) -> object:
            # bind explícito por default-arg (B023): la closure se consume en
            # este mismo ciclo, pero el bind elimina la clase entera de bug.
            idx = mapa.get(clave)
            return _celdas[idx] if idx is not None and idx < len(_celdas) else None

        try:
            prefijo = str(celda("prefijo") or "").strip()
            folio = str(celda("folio") or "").strip()
            if not folio:
                raise FilaIlegible(f"fila {n}: sin folio/número de documento")
            cufe = str(celda("cufe") or "").strip()
            if not cufe:
                raise FilaIlegible(f"fila {n}: sin CUFE")
            filas.append(
                {
                    "fila": n,
                    "tipo_documento": str(celda("tipo_documento") or "").strip(),
                    "numero": f"{prefijo}{folio}",
                    "cufe": cufe,
                    "fecha": _fecha_iso(celda("fecha"), n),
                    "nit_emisor": str(celda("nit_emisor") or "").strip(),
                    "nombre_emisor": str(celda("nombre_emisor") or "").strip(),
                    "iva": _monto(celda("iva"), "IVA", n),
                    "total": _monto(celda("total"), "Total", n),
                    "nit_receptor": str(celda("nit_receptor") or "").strip(),
                    "inc": _monto_opcional(celda("inc"), "INC", n),
                    "bolsas": _monto_opcional(celda("bolsas"), "INC Bolsas", n),
                    "rete_iva": _monto_opcional(celda("rete_iva"), "Rete IVA", n),
                    "rete_fuente": _monto_opcional(
                        celda("rete_fuente"), "Rete Renta", n
                    ),
                    "rete_ica": _monto_opcional(celda("rete_ica"), "Rete ICA", n),
                }
            )
        except FilaIlegible as e:
            filas.append({"fila": n, "error": str(e)})
    return filas


def campos_desde_fila(fila: dict, *, nits_auteco: frozenset[str]) -> dict:
    """Fila cruda válida → campos del Document Factura (compra recibida).

    Misma semántica de deducibilidad que `ingesta.campos_desde_dian`: Auteco por
    NIT → deducible/decidida/origen auteco (factura con VARIOS NITs — config
    {"nits": [...]}, CEO 2026-08-11); el resto sin decidir (el operador marca
    después — contador del §2). El Excel de la DIAN no trae total bruto ni
    base gravada por línea → None (R5: no se inventa)."""
    es_auteco = fila["nit_emisor"] in nits_auteco
    nombre = fila["nombre_emisor"] or f"NIT {fila['nit_emisor']}"
    return {
        "tipo": TipoFactura.compra,
        "origen": OrigenFactura.auteco if es_auteco else OrigenFactura.sin_clasificar,
        "numero": fila["numero"],
        "tercero_nombre": nombre,
        "tercero_nit": fila["nit_emisor"],
        "fecha": fila["fecha"],
        "base_gravable": None,
        "total_bruto": None,
        "tarifa_iva": None,
        "iva_valor": fila["iva"],
        "total": fila["total"],
        "deducible": es_auteco,
        "deducible_decidido": es_auteco,
        "cufe": fila["cufe"],
        "tipo_documento": fila["tipo_documento"],
        "signo": 1,
        # el export no trae tipo de contribuyente → None = se trata como PII
        # por precaución (A17), igual que la captura manual.
        "tipo_contribuyente": None,
        "inc_valor": fila.get("inc", Decimal("0.00")),
        "bolsas": fila.get("bolsas", Decimal("0.00")),
        "otros_impuestos": Decimal("0.00"),
        "rete_fuente": fila.get("rete_fuente", Decimal("0.00")),
        "rete_iva": fila.get("rete_iva", Decimal("0.00")),
        "rete_ica": fila.get("rete_ica", Decimal("0.00")),
    }


def es_tipo_soportado(tipo_documento: str) -> bool:
    n = _norm(tipo_documento)
    return n.startswith(_TIPO_FACTURA_PREFIJO) and "nota" not in n
