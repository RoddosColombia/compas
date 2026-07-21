# backend/app/parsers/bank_parsers.py
"""Parsers de extractos bancarios — Bancolombia, BBVA, Global66 (Spec §1.5).

Portados en espíritu de SISMO v2 (`services/bank_parsers.py`) y adaptados a las
reglas innegociables de COMPAS:

  - **Regla 1 (Decimal, nunca float):** los montos se construyen como `Decimal`
    vía el tipo `Money`, que rechaza `float`. Los números de openpyxl (que llegan
    como `float`) se convierten con `Decimal(str(v))` para no arrastrar ruido binario.
  - **Regla 7 (transforma, no interpreta):** una fila con fecha o monto no
    parseables NO se adivina ni se traga en silencio: se acumula en
    `ResultadoParseo.errores` con el número de fila. Solo las filas totalmente
    vacías (o de valor 0, que no son movimiento de caja) se omiten sin error.
  - **Regla 7 (Global66):** se conservan `moneda_original` y `tasa_cambio`. Hoy
    solo se lee la hoja 'Movimientos de cuenta COP' → `moneda_original='COP'`,
    `tasa_cambio=1` (es un hecho de esa hoja, no una interpretación). El mapeo FX
    multi-moneda real requiere un export Global66 de muestra (TODO Sprint 1).

Formatos (heredados de la realidad de cada banco):
  - Bancolombia: .xlsx, hoja 'Extracto', headers fila 15, fecha d/m sin año.
  - BBVA: .xlsx, hoja activa, headers fila 14, fecha d-m-Y.
  - Global66: .xls/.xlsx, hoja 'Movimientos de cuenta COP', headers fila 4.
"""

import os
import shutil
import tempfile
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum

import openpyxl
from pydantic import BaseModel, ConfigDict

from app.core.money import Money
from app.core.time import today_bogota
from app.domain.bancos import Banco


class TipoMovimiento(StrEnum):
    DEBITO = "debito"
    CREDITO = "credito"


class MovimientoBancario(BaseModel):
    """Un movimiento parseado, normalizado y con dinero en Decimal (regla 1)."""

    model_config = ConfigDict(strict=True, extra="forbid")

    fecha: date
    descripcion: str
    monto: Money  # positivo (magnitud); el signo vive en `tipo`
    tipo: TipoMovimiento
    banco: Banco
    # Global66 (regla 7): moneda de origen y tasa aplicada.
    moneda_original: str | None = None
    tasa_cambio: Decimal | None = None
    referencia: str | None = None


class ErrorFila(BaseModel):
    """Fila que el parser no pudo transformar sin adivinar (regla 7)."""

    fila: int
    motivo: str
    valor_crudo: str | None = None


class ResultadoParseo(BaseModel):
    banco: Banco
    movimientos: list[MovimientoBancario]
    errores: list[ErrorFila]


class _FilaError(Exception):
    """Interno: se eleva por fila y se convierte en `ErrorFila`."""

    def __init__(self, motivo: str, valor_crudo: str | None = None) -> None:
        super().__init__(motivo)
        self.motivo = motivo
        self.valor_crudo = valor_crudo


# ── Utilidades ───────────────────────────────────────────────────────────


# F-22 (Kimi M-2): topes duros ANTES de gastar CPU/memoria en un archivo hostil.
MAX_FILAS = 20_000
_MAX_DESCOMPRIMIDO = 200 * 1024 * 1024  # xlsx = zip; 200 MB descomprimidos
_MAX_RATIO = 100  # ratio descomprimido/comprimido típico de una zip-bomb: >1000


def _validar_zip(file_path: str) -> None:
    """F-22: acota el ratio de descompresión (un .xlsx es un zip; una bomba de
    10 MB puede expandir a GB). Lanza ValueError antes de abrir con openpyxl."""
    try:
        comprimido = os.path.getsize(file_path)
        with zipfile.ZipFile(file_path) as z:
            total = sum(i.file_size for i in z.infolist())
    except zipfile.BadZipFile:
        return  # no es zip (p. ej. .xls binario legacy): openpyxl decidirá
    ratio_excedido = comprimido > 0 and total / comprimido > _MAX_RATIO
    if total > _MAX_DESCOMPRIMIDO or ratio_excedido:
        raise ValueError(
            "el extracto excede el límite de descompresión permitido (F-22): "
            f"{total // (1024 * 1024)} MB descomprimidos"
        )


def _open_workbook(file_path: str):
    """Abre el .xlsx; si la extensión .xls confunde a openpyxl, copia a temp."""
    _validar_zip(file_path)  # F-22 (Kimi M-2)
    try:
        return openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        tmp = os.path.join(tempfile.mkdtemp(), "extract.xlsx")
        shutil.copy2(file_path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True)


def _check_tope_filas(fila_datos: int) -> None:
    """F-22: tope de filas de datos (Kimi M-2). Error explícito, no minutos de CPU."""
    if fila_datos > MAX_FILAS:
        raise ValueError(
            f"el extracto supera el tope de {MAX_FILAS} filas de datos (F-22)"
        )


def _cell(row: tuple, idx: int):
    return row[idx] if idx is not None and idx < len(row) else None


def _fila_vacia(row: tuple) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in row)


def _mapear_columnas(ws, header_row: int, spec: dict[str, str]) -> dict[str, int]:
    """Ubica columnas por substring del header (case-insensitive)."""
    headers = [str(c.value or "").strip().upper() for c in ws[header_row]]
    col: dict[str, int] = {}
    for key, needle in spec.items():
        for i, h in enumerate(headers):
            if needle in h:
                col[key] = i
                break
    faltan = [k for k in spec if k not in col]
    if faltan:
        raise ValueError(
            f"Columnas no encontradas en la fila {header_row}: {faltan}. "
            f"Headers vistos: {headers}"
        )
    return col


def _a_decimal(valor) -> Decimal:
    """Convierte un valor de celda a Decimal con signo. Eleva `_FilaError` si es
    ambiguo (regla 7: no se adivina 0)."""
    if valor is None:
        raise _FilaError("monto vacío")
    if isinstance(valor, bool):  # bool es subclase de int
        raise _FilaError("monto booleano inválido", str(valor))
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    s = str(valor).strip().replace("$", "").replace(" ", "")
    if s == "":
        raise _FilaError("monto vacío", str(valor))
    negativo = s.startswith("-")
    s = s.lstrip("+-").replace(",", "")  # coma = separador de miles
    try:
        d = Decimal(s)
    except InvalidOperation:
        raise _FilaError("monto no numérico", str(valor)) from None
    return -d if negativo else d


def _fecha_dmy(raw, con_anio: bool) -> date:
    """Parsea fecha. `con_anio=False` → formato d/m sin año (Bancolombia): se
    completa con el año actual (Bogotá)."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if con_anio:
        try:
            return datetime.strptime(s, "%d-%m-%Y").date()
        except (ValueError, TypeError):
            raise _FilaError("fecha inválida", s) from None
    # d/m/Y explícito (con año) → tal cual.
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except (ValueError, TypeError):
        pass
    # d/m sin año: se completa con el año actual, PERO una fecha d/m no puede ser
    # futura → si cae en el futuro (frontera dic/ene: cargar el 2-ene un movimiento
    # del 31-dic), es del año anterior (Kimi M-01). El fix definitivo es leer el año
    # del encabezado del extracto; se hará al congelar los fixtures reales (S1-01).
    hoy = today_bogota()
    try:
        dt = datetime.strptime(f"{s}/{hoy.year}", "%d/%m/%Y").date()
    except (ValueError, TypeError):
        raise _FilaError("fecha inválida", s) from None
    if dt > hoy:
        try:
            dt = dt.replace(year=dt.year - 1)
        except ValueError:  # 29-feb en año no bisiesto
            dt = dt.replace(year=dt.year - 1, day=28)
    return dt


def _fecha_iso(raw) -> date:
    """Global66: 'YYYY-MM-DD HH:MM:SS' (o datetime)."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    raise _FilaError("fecha inválida", s)


# ── Detección ──────────────────────────────────────────────────────────────


def detectar_banco(file_path: str) -> Banco:
    """Auto-detecta el banco. Lanza ValueError en formato/estructura no soportada."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise ValueError(
            f"Formato no soportado: {ext or '(sin extensión)'}. COMPAS procesa "
            ".xlsx/.xls (Bancolombia, BBVA, Global66)."
        )
    wb = _open_workbook(file_path)
    try:
        nombres = wb.sheetnames
        if any("Movimientos de cuenta COP" in n for n in nombres):
            return Banco.GLOBAL66
        if "Extracto" in nombres:
            return Banco.BANCOLOMBIA
        ws = wb.active
        fila14 = [str(c.value or "") for c in ws[14]]
        if any("FECHA DE OPERACI" in c.upper() for c in fila14):
            return Banco.BBVA
        raise ValueError("No se pudo identificar el banco (Bancolombia/BBVA/Global66).")
    finally:
        wb.close()


# ── Parsers por banco ────────────────────────────────────────────────────


def _parse_signo(
    file_path: str, banco: Banco, hoja: str | None, header_row: int, con_anio: bool
) -> ResultadoParseo:
    """Común a Bancolombia/BBVA: FECHA/DESCRIPCIÓN/VALOR con signo en el monto."""
    wb = _open_workbook(file_path)
    movimientos: list[MovimientoBancario] = []
    errores: list[ErrorFila] = []
    try:
        ws = wb[hoja] if hoja else wb.active
        col = _mapear_columnas(
            ws,
            header_row,
            {
                "fecha": "FECHA",
                "descripcion": "CONCEPTO" if banco is Banco.BBVA else "DESCRIPCI",
                "valor": "IMPORTE" if banco is Banco.BBVA else "VALOR",
            },
        )
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
        ):
            _check_tope_filas(r_idx - header_row)  # F-22 (Kimi M-2)
            if _fila_vacia(row):
                continue
            try:
                fecha = _fecha_dmy(_cell(row, col["fecha"]), con_anio)
                signed = _a_decimal(_cell(row, col["valor"]))
                if signed == 0:
                    continue
                mov = MovimientoBancario(
                    fecha=fecha,
                    descripcion=str(_cell(row, col["descripcion"]) or "").strip(),
                    monto=abs(signed),
                    tipo=(
                        TipoMovimiento.DEBITO if signed < 0 else TipoMovimiento.CREDITO
                    ),
                    banco=banco,
                )
            except _FilaError as e:
                errores.append(
                    ErrorFila(fila=r_idx, motivo=e.motivo, valor_crudo=e.valor_crudo)
                )
                continue
            movimientos.append(mov)
    finally:
        wb.close()
    return ResultadoParseo(banco=banco, movimientos=movimientos, errores=errores)


def parse_bancolombia(file_path: str) -> ResultadoParseo:
    return _parse_signo(file_path, Banco.BANCOLOMBIA, "Extracto", 15, con_anio=False)


def parse_bbva(file_path: str) -> ResultadoParseo:
    return _parse_signo(file_path, Banco.BBVA, None, 14, con_anio=True)


def parse_global66(file_path: str) -> ResultadoParseo:
    """Global66: hoja 'Movimientos de cuenta COP'. Débito en col C, crédito en D."""
    wb = _open_workbook(file_path)
    movimientos: list[MovimientoBancario] = []
    errores: list[ErrorFila] = []
    try:
        ws = None
        for nombre in wb.sheetnames:
            if "Movimientos de cuenta COP" in nombre:
                ws = wb[nombre]
                break
        if ws is None:
            raise ValueError("No se encontró la hoja 'Movimientos de cuenta COP'.")
        for r_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            _check_tope_filas(r_idx - 4)  # F-22 (Kimi M-2)
            if _fila_vacia(row):
                continue
            debito, credito = _cell(row, 2), _cell(row, 3)
            tiene_deb = debito not in (None, "", 0)
            tiene_cred = credito not in (None, "", 0)
            if not tiene_deb and not tiene_cred:
                continue  # GMF informativo / fila sin valor
            try:
                fecha = _fecha_iso(_cell(row, 1))
                if tiene_deb:
                    monto, tipo = abs(_a_decimal(debito)), TipoMovimiento.DEBITO
                else:
                    monto, tipo = abs(_a_decimal(credito)), TipoMovimiento.CREDITO
                if monto == 0:
                    continue
                partes = [
                    p
                    for p in (
                        str(_cell(row, 0) or "").strip(),
                        str(_cell(row, 13) or "").strip(),
                        str(_cell(row, 7) or "").strip(),
                    )
                    if p
                ]
                ref = str(_cell(row, 12) or "").strip() or None
                mov = MovimientoBancario(
                    fecha=fecha,
                    descripcion=" — ".join(partes),
                    monto=monto,
                    tipo=tipo,
                    banco=Banco.GLOBAL66,
                    moneda_original="COP",
                    tasa_cambio=Decimal("1"),
                    referencia=ref,
                )
            except _FilaError as e:
                errores.append(
                    ErrorFila(fila=r_idx, motivo=e.motivo, valor_crudo=e.valor_crudo)
                )
                continue
            movimientos.append(mov)
    finally:
        wb.close()
    return ResultadoParseo(
        banco=Banco.GLOBAL66, movimientos=movimientos, errores=errores
    )


def parse_extracto(file_path: str, banco: Banco | None = None) -> ResultadoParseo:
    """Punto de entrada: auto-detecta el banco (si no se pasa) y rutea."""
    if banco is None:
        banco = detectar_banco(file_path)
    if banco is Banco.BANCOLOMBIA:
        return parse_bancolombia(file_path)
    if banco is Banco.BBVA:
        return parse_bbva(file_path)
    if banco is Banco.GLOBAL66:
        return parse_global66(file_path)
    raise ValueError(f"Banco no soportado: {banco}")
