# backend/app/cartera_previa/cronograma.py
"""SUP-4 (CEO 2026-08-22) — el cronograma real de pagos, agregado para el motor.

"Semanalmente, lunes podría ser, se carga el cronograma y el loantape para que
actualice info" + "cargar tal cantidad de detalle volverá más pesada la app y no es
necesario; necesito datos completos que tú puedes calcular".

Así que este parser **no persiste las ~9.900 cuotas**: las agrega a dos cosas ligeras
que COMPAS ya sabe consumir.

  1. `serie` — recaudo pendiente y nº de créditos pagando POR SEMANA GLOBAL (el ancla
     del motor: miércoles 2026-03-04 = semana 1). Es la cartera ya originada: cada
     crédito con SU cuota real pactada, incluidos plazos que ni están en el catálogo
     (P39S). Lo que el motor proyecta con las cuotas nuevas es solo lo FUTURO.
  2. `colocaciones_por_mes` — la cuota 0 es el desembolso, así que marca el mes en que
     se colocó cada moto. Con eso `rampa_mes_en_curso` deja el mes vivo en el
     REMANENTE hacia la meta (criterio CEO: agosto vive con la meta de 70 y se cierra
     con lo realmente logrado).

Criterios de dinero: las cuotas PAGADAS no se proyectan; las PARCIALES cuentan solo su
saldo; lo VENCIDO sin pagar se reporta aparte (es mora real medida — no se proyecta ni
se inventa cuándo entra). Regla 7: encabezados que no cuadran → error que LISTA
esperado vs encontrado; una fila ilegible no frena el lote, se reporta.
"""

import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

# Mismo ancla que `motor.ANCLA_SEMANA` (miércoles 2026-03-04 = semana 1). Se replica
# aquí para que el parser sea PURO (sin importar el motor).
ANCLA_SEMANA = date(2026, 3, 4)
MAX_FILAS_BUSQUEDA_ENCABEZADO = 10

# clave interna → alias aceptados (normalizados: minúsculas y sin tildes)
COLUMNAS: dict[str, tuple[str, ...]] = {
    # el PRIMER alias es la etiqueta legible que se muestra en el error (regla 7)
    "credito": ("crédito", "credito", "loanbook_codigo", "codigo"),
    "cuota": ("cuota #", "cuota", "n cuota", "numero cuota"),
    "fecha": ("fecha programada", "fecha", "fecha_programada"),
    "monto": ("monto total", "monto", "valor cuota"),
    "estado": ("estado",),
}
OPCIONALES: dict[str, tuple[str, ...]] = {
    "pagado": ("pagado",),
    "saldo": ("saldo",),
}


class EncabezadosNoReconocidos(Exception):
    """El archivo no trae las columnas del contrato (regla 7: fail-loud)."""


class FilaIlegible(Exception):
    """Una fila puntual no se pudo transformar sin interpretar."""


@dataclass(frozen=True)
class ResumenCronograma:
    """Lo que el cronograma aporta, ya digerido."""

    serie: list[dict]  # [{semana_global, recaudo, n_activos}] — para el motor
    colocaciones_por_mes: dict[str, int]  # 'YYYY-MM' → motos colocadas (cuota 0)
    creditos: int
    cuotas_futuras: int
    recaudo_futuro: Decimal
    vencido_sin_pagar: Decimal  # mora real medida (NO se proyecta)
    creditos_en_mora: int
    errores: list[str] = field(default_factory=list)


def _norm(v: object) -> str:
    s = str(v or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _mapear(celdas: list[object]) -> dict[str, int] | None:
    normalizadas = [_norm(c) for c in celdas]
    mapa: dict[str, int] = {}
    for clave, alias in {**COLUMNAS, **OPCIONALES}.items():
        alias_norm = {_norm(a) for a in alias}
        for i, h in enumerate(normalizadas):
            if h and h in alias_norm:
                mapa[clave] = i
                break
    return mapa if all(k in mapa for k in COLUMNAS) else None


def _fecha(v: object, fila: int) -> date:
    if isinstance(v, date):
        return v
    if hasattr(v, "date"):
        return v.date()
    s = str(v or "").strip()[:10]
    for sep in ("-", "/"):
        partes = s.split(sep)
        if len(partes) == 3:
            try:
                a, b, c = (int(p) for p in partes)
                return date(a, b, c) if len(partes[0]) == 4 else date(c, b, a)
            except ValueError:
                break
    raise FilaIlegible(f"fila {fila}: fecha programada ilegible ('{s}')")


def _monto(v: object, fila: int, campo: str) -> Decimal:
    if v in (None, ""):
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace("$", "").replace(" ", "")
    try:
        if "," in s and "." in s:  # es-CO: 1.452,94
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return Decimal(s)
    except InvalidOperation:
        raise FilaIlegible(f"fila {fila}: {campo} ilegible ('{v}')") from None


def semana_global(f: date) -> int:
    """Réplica de `motor.indice_semana`: floor((fecha − ancla)/7) + 1."""
    return (f - ANCLA_SEMANA).days // 7 + 1


def parsear_cronograma(contenido: bytes, hoy: date) -> ResumenCronograma:
    """xlsx del cronograma → series agregadas. `hoy` marca el corte entre lo vencido
    (mora medida) y lo futuro (lo que el motor debe esperar)."""
    wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active

    mapa: dict[str, int] | None = None
    fila_encabezado = 0
    mejores: list[str] = []
    for i, row in enumerate(ws.iter_rows(max_row=MAX_FILAS_BUSQUEDA_ENCABEZADO), 1):
        celdas = [c.value for c in row]
        mapa = _mapear(celdas)
        if mapa is not None:
            fila_encabezado = i
            break
        con_texto = [str(c) for c in celdas if c not in (None, "")]
        if len(con_texto) > len(mejores):
            mejores = con_texto
    if mapa is None:
        esperadas = ", ".join(a[0] for a in COLUMNAS.values())
        encontradas = ", ".join(mejores) or "(ninguna)"
        raise EncabezadosNoReconocidos(
            f"encabezados no reconocidos: se esperaba una fila con [{esperadas}]; "
            f"lo más parecido fue [{encontradas}]. Verifica que sea el export del "
            "'Cronograma General'."
        )

    recaudo: dict[int, Decimal] = {}
    activos: dict[int, set] = {}
    colocaciones: dict[str, int] = {}
    creditos: set[str] = set()
    en_mora: set[str] = set()
    vencido = Decimal("0")
    cuotas_futuras = 0
    errores: list[str] = []

    for n, row in enumerate(
        ws.iter_rows(min_row=fila_encabezado + 1), fila_encabezado + 1
    ):
        celdas = [c.value for c in row]
        if all(c in (None, "") for c in celdas):
            continue

        def celda(clave: str, _c: list[object] = celdas) -> object:
            i = mapa.get(clave)
            return _c[i] if i is not None and i < len(_c) else None

        try:
            credito = str(celda("credito") or "").strip()
            if not credito:
                continue
            creditos.add(credito)
            f = _fecha(celda("fecha"), n)
            estado = _norm(celda("estado"))
            n_cuota = celda("cuota")
            # la cuota 0 es el DESEMBOLSO: marca el mes de colocación, no recaudo
            if str(n_cuota).strip() in ("0", "0.0"):
                clave_mes = f"{f.year:04d}-{f.month:02d}"
                colocaciones[clave_mes] = colocaciones.get(clave_mes, 0) + 1
                continue
            if estado == "pagada":
                continue
            # lo que FALTA de la cuota: el saldo manda (las parciales ya abonaron)
            saldo = celda("saldo")
            falta = (
                _monto(saldo, n, "saldo")
                if saldo not in (None, "")
                else _monto(celda("monto"), n, "monto total")
            )
            if falta <= 0:
                continue
            if f < hoy:
                vencido += falta
                en_mora.add(credito)
                continue
            s = semana_global(f)
            recaudo[s] = recaudo.get(s, Decimal("0")) + falta
            activos.setdefault(s, set()).add(credito)
            cuotas_futuras += 1
        except FilaIlegible as e:
            errores.append(str(e))

    serie = [
        {
            "semana_global": s,
            "recaudo": recaudo[s].quantize(Decimal("0.01")),
            "n_activos": len(activos[s]),
        }
        for s in sorted(recaudo)
    ]
    return ResumenCronograma(
        serie=serie,
        colocaciones_por_mes=colocaciones,
        creditos=len(creditos),
        cuotas_futuras=cuotas_futuras,
        recaudo_futuro=sum((f["recaudo"] for f in serie), Decimal("0")),
        vencido_sin_pagar=vencido,
        creditos_en_mora=len(en_mora),
        errores=errores,
    )


def rampa_mes_en_curso(
    colocaciones_por_mes: dict[str, int], mes: tuple[int, int], meta: int
) -> dict[str, int]:
    """El mes VIVO proyecta solo el REMANENTE hacia la meta (criterio CEO 2026-08-22).

    Las motos ya colocadas recaudan por la serie con su cuota real; las que faltan las
    proyecta el motor con la cuota nueva. Si ya se superó la meta, el remanente es 0 y
    manda la realidad (nunca resta)."""
    clave = f"{mes[0]:04d}-{mes[1]:02d}"
    return {clave: max(0, meta - colocaciones_por_mes.get(clave, 0))}
