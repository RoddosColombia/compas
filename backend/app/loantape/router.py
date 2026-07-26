# backend/app/loantape/router.py
"""/api/v1/loantape — carga del LoanTape semanal de SISMO-V3 + aging por tramo.

RBAC: carga con `cargas:gestionar` (+ verify_origin); aging con `dashboard:leer`. El
upload acepta CSV o Excel (.xlsx) según el contrato; el parser transforma, no adivina
(fila ambigua → 422). Sin Idempotency-Key: el upsert por (credito_id, fecha_corte) hace
inocuo el replay (pisa el corte). Montos como string en la respuesta (regla 1)."""

import csv
import io

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from openpyxl import load_workbook

from app.auth.deps import require_permission
from app.auth.models import User
from app.auth.router import verify_origin
from app.core.money import money_str
from app.loantape import service

router = APIRouter(prefix="/loantape", tags=["loantape"])

_MAX_BYTES = 20 * 1024 * 1024  # 20 MB (tope defensivo)


def _filas_csv(contenido: bytes) -> list[dict]:
    texto = contenido.decode("utf-8-sig")  # tolera BOM
    return list(csv.DictReader(io.StringIO(texto)))


def _filas_xlsx(contenido: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas = ws.iter_rows(values_only=True)
    try:
        encabezado = [str(c).strip() if c is not None else "" for c in next(filas)]
    except StopIteration:
        return []
    out: list[dict] = []
    for fila in filas:
        if all(c is None for c in fila):
            continue
        out.append(
            {
                encabezado[i]: ("" if v is None else str(v))
                for i, v in enumerate(fila)
                if i < len(encabezado)
            }
        )
    return out


@router.post("/carga", status_code=201)
async def cargar(
    archivo: UploadFile,
    user: User = Depends(require_permission("cargas:gestionar")),
    _: None = Depends(verify_origin),
):
    nombre = (archivo.filename or "").lower()
    contenido = await archivo.read(_MAX_BYTES + 1)
    if len(contenido) > _MAX_BYTES:
        raise HTTPException(413, "el archivo supera el tope de 20 MB")
    if nombre.endswith(".csv"):
        filas = _filas_csv(contenido)
    elif nombre.endswith(".xlsx"):
        filas = _filas_xlsx(contenido)
    else:
        raise HTTPException(422, "el archivo debe ser .csv o .xlsx")
    try:
        n = await service.cargar_loantape(filas, usuario_id=user.id)
    except service.LoanTapeError as e:
        raise HTTPException(e.status, e.detalle) from e
    return {"cargados": n}


@router.get("/aging")
async def aging(
    fecha_corte: str | None = Query(default=None),
    _: User = Depends(require_permission("dashboard:leer")),
):
    data = await service.obtener_aging(fecha_corte)
    return {
        "fecha_corte": data["fecha_corte"],
        "tramos": [
            {
                "tramo": t["tramo"],
                "etiqueta": t["etiqueta"],
                "n_creditos": t["n_creditos"],
                "saldo_en_mora": money_str(t["saldo_en_mora"]),
            }
            for t in data["tramos"]
        ],
    }
