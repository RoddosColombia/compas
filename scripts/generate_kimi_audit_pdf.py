#!/usr/bin/env python
"""Genera el PDF de una SOLICITUD DE AUDITORÍA para Kimi (COMPAS).

Procedimiento portado de SISMO-V3 (ver CLAUDE.md → «Auditoría adversarial con Kimi»).
Kimi no tiene CLI/API aquí: este PDF se sube a mano al chat de Kimi y la respuesta se
pega en el AUDITORIA-KIMI-*.md correspondiente.

Uso:
    python scripts/generate_kimi_audit_pdf.py <SOLICITUD.md> [más_docs.md ...] [salida.pdf]

Se puede pasar más de un .md (p. ej. la SOLICITUD + la EVIDENCIA):
se renderizan en orden, con salto de página entre cada uno. Si el último argumento
termina en .pdf, es la ruta de salida; si no, se usa el default.

Junta:
  1. El texto de los .md indicados (markdown → texto formateado).
  2. Un extracto del control (tracker docs/COMPAS_Control_Desarrollo.xlsx): Tareas, DoD, Gates.
  3. La rúbrica del umbral (≥ 9.0).

Salida por defecto: PAQUETE.pdf en la MISMA carpeta del primer .md (la carpeta de
la ronda: planning/phases/<fase>/auditorias/<RONDA>/PAQUETE.pdf).
"""

from __future__ import annotations

import sys
from pathlib import Path

from fpdf import FPDF
from fpdf.enums import XPos, YPos
from openpyxl import load_workbook

# fpdf2 2.8: por defecto multi_cell deja el cursor a la derecha; forzamos salto
# de línea real (vuelve al margen izquierdo en la siguiente fila).
_NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}

REPO_ROOT = Path(__file__).resolve().parents[1]
TRACKER = REPO_ROOT / "docs" / "COMPAS_Control_Desarrollo.xlsx"

# fpdf2 con fuentes core codifica latin-1. Mapear lo que no es latin-1 a ASCII.
_REPLACEMENTS = {
    "—": "-", "–": "-", "→": "->", "←": "<-", "≥": ">=", "≤": "<=",
    "•": "-", "·": "-", "…": "...", "“": '"', "”": '"', "‘": "'", "’": "'",
    "✅": "[OK]", "❌": "[X]", "⚠": "[!]", "🟢": "[VER]", "🟡": "[AMA]",
    "🔴": "[ROJ]", "✔": "[OK]", "✖": "[X]", "≡": "=", "×": "x",
}


def sanitize(s: str) -> str:
    for k, v in _REPLACEMENTS.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def _fit(pdf: FPDF, text: str) -> str:
    """Parte por ancho medido cualquier palabra más ancha que el área útil de
    la página (fpdf2 lanza si un token no cabe). Llamar tras set_font."""
    epw = pdf.epw
    words: list[str] = []
    for word in text.split(" "):
        if not word or pdf.get_string_width(word) <= epw:
            words.append(word)
            continue
        cur = ""
        for ch in word:
            if pdf.get_string_width(cur + ch) > epw and cur:
                words.append(cur)
                cur = ch
            else:
                cur += ch
        if cur:
            words.append(cur)
    return " ".join(words)


class PDF(FPDF):
    def header(self) -> None:
        self.set_font("helvetica", "I", 8)
        self.set_text_color(120)
        self.cell(0, 6, "COMPAS - Solicitud de auditoria Kimi", align="R")
        self.ln(4)
        self.set_text_color(0)


def _write_markdown(pdf: PDF, text: str) -> None:
    for raw in text.splitlines():
        line = sanitize(raw.rstrip())
        if not line:
            pdf.ln(3)
            continue
        if line.startswith("### "):
            pdf.set_font("helvetica", "B", 11)
            pdf.multi_cell(0, 6, _fit(pdf, line[4:]), **_NL)
        elif line.startswith("## "):
            pdf.set_font("helvetica", "B", 13)
            pdf.ln(1)
            pdf.multi_cell(0, 7, _fit(pdf, line[3:]), **_NL)
        elif line.startswith("# "):
            pdf.set_font("helvetica", "B", 15)
            pdf.multi_cell(0, 8, _fit(pdf, line[2:]), **_NL)
        else:
            pdf.set_font("helvetica", "", 10)
            pdf.multi_cell(0, 5, _fit(pdf, line), **_NL)


def _append_tracker_extract(pdf: PDF) -> None:
    if not TRACKER.exists():
        return
    wb = load_workbook(TRACKER, data_only=True)
    pdf.add_page()
    pdf.set_font("helvetica", "B", 15)
    pdf.multi_cell(0, 8, "Extracto del control (tracker)", **_NL)
    pdf.ln(2)
    for sheet in ("Tareas", "DoD", "Gates"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        pdf.set_font("helvetica", "B", 12)
        pdf.multi_cell(0, 6, f"Hoja: {sheet}", **_NL)
        pdf.set_font("helvetica", "", 8)
        for row in ws.iter_rows(values_only=True):
            cells = [sanitize(str(c)) for c in row if c is not None]
            if cells:
                pdf.multi_cell(0, 4, _fit(pdf, " | ".join(cells)), **_NL)
        pdf.ln(3)


def main() -> None:
    args = sys.argv[1:]
    if not args:
        sys.exit("Uso: python scripts/generate_kimi_audit_pdf.py <SOLICITUD.md> [más.md ...] [salida.pdf]")

    out_arg = args[-1] if args[-1].lower().endswith(".pdf") else None
    md_args = args[:-1] if out_arg else args
    md_files = [Path(a).resolve() for a in md_args]
    for f in md_files:
        if not f.exists():
            sys.exit(f"No existe: {f}")
    if not md_files:
        sys.exit("Falta al menos un .md de entrada.")

    # Por defecto el PDF se llama PAQUETE.pdf y vive JUNTO a la SOLICITUD, en la
    # carpeta de la ronda (planning/phases/<fase>/auditorias/<RONDA>/). Así cada
    # intercambio con Kimi queda autocontenido y no hay nombres ad-hoc.
    out = Path(out_arg) if out_arg else md_files[0].parent / "PAQUETE.pdf"
    out.parent.mkdir(parents=True, exist_ok=True)

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    for i, md in enumerate(md_files):
        pdf.add_page()
        _write_markdown(pdf, md.read_text(encoding="utf-8"))
    _append_tracker_extract(pdf)

    # Rúbrica
    pdf.add_page()
    pdf.set_font("helvetica", "B", 13)
    pdf.multi_cell(0, 7, "Rubrica", **_NL)
    pdf.set_font("helvetica", "", 10)
    pdf.multi_cell(
        0, 5,
        _fit(pdf, sanitize(
            "Umbral de aprobacion: >= 9.0 (plan y codigo). "
            "Kimi es auditor adversarial: no genera codigo. "
            "Auditar con lupa los 'puntos a auditar', el cumplimiento de las reglas "
            "innegociables de CLAUDE.md y del DoD (Spec 5). "
            "Veredicto: APROBADO (>= 9.0, sin P0/P1) o RECHAZADO con hallazgos accionables."
        )),
    )

    pdf.output(str(out))
    print(f"PDF generado: {out}")


if __name__ == "__main__":
    main()
