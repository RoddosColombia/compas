#!/usr/bin/env python
"""Espejo de agosto 2026 (Global66) en PROD — dejar PROD agosto = Excel "Global66 ago-2026".

CONTEXTO (verificado read-only 2026-08-13/14): en PROD solo estaban cargados ago 1-4
(42 tx); la fuente al día — la hoja clasificada del libro maestro — trae ago 1-12 (~225
mov). Esta migración inserta los faltantes y ALINEA el rubro de todos al que dicta el
Excel (columna *Detalle*), para que la ejecución del presupuesto de agosto sea fiel.

Distinto de la carga inicial (`20260726_carga_inicial_global66.py`, que clasifica por
reglas): aquí el rubro sale EXACTO de la columna *Detalle* del Excel (fidelidad a la
clasificación del CEO), no de las reglas. La ingesta reusa `parse_global66` verbatim
(mismo `id_banco`/ocurrencia → dedup exacta) y el patrón transaccional de `procesar_carga`
(regla 8). Las reclasificaciones de lo ya cargado usan `reclasificar_transaccion`.

Idempotente: dedup por (banco, id_banco) no re-inserta; la reclasificación se salta lo ya
correcto; el rubro nuevo se crea solo si falta.

GATE (CLAUDE.md): migración de datos reales / carga bancaria → dry-run + auditoría Kimi
≥ 9.0 + GO del CEO antes de `--commit`. Correr `--commit` contra prod es el acto gated.

Uso (Windows: PYTHONUTF8=1). La URI NUNCA va por argv — se lee de la env var:
    export MONGODB_URI_COMPAS="mongodb+srv://..."
    PYTHONUTF8=1 python migrations/20260814_espejo_agosto_global66.py            # dry-run
    PYTHONUTF8=1 python migrations/20260814_espejo_agosto_global66.py --commit    # escribe
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
import unicodedata
import warnings
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, "backend")

import openpyxl  # noqa: E402
from beanie import PydanticObjectId  # noqa: E402

from app.audit import service as audit_service  # noqa: E402
from app.audit.events import AuditEvento  # noqa: E402
from app.audit.service import emit_audit  # noqa: E402
from app.cargas.service import _clave_ocurrencia  # noqa: E402
from app.cargas.mapper import movimiento_a_transaccion  # noqa: E402
from app.core.time import now_utc  # noqa: E402
from app.db import mongo  # noqa: E402
from app.domain.bancos import Banco  # noqa: E402
from app.domain.mes_control import EstadoMes, MesControl  # noqa: E402
from app.domain.rubro import Rubro, TipoFlujo, es_rubro_clasificable  # noqa: E402
from app.domain.transaccion import Transaccion  # noqa: E402
from app.parsers.bank_parsers import TipoMovimiento, parse_global66  # noqa: E402

warnings.simplefilter("ignore")

MES = "2026-08-01"
SNAPSHOT = "docs/modelo/Global66_ago2026_clasificado.xlsx"
HOJA = "Movimientos de cuenta COP"

# Control fail-loud: Σ del Excel (footer de la hoja). Mismatch → aborta sin escribir.
SIGMA_EGRESOS = Decimal("150673115.59")
SIGMA_INGRESOS = Decimal("99424130.75")

RUBRO_RENDIMIENTOS = "Rendimientos bancarios"

# Categorías (Detalle) que NO calzan literal con el nombre del rubro (decisiones CEO).
# El resto mapea directo: categoria == nombre del rubro.
ALIAS_CATEGORIA_RUBRO = {
    "operativo": "Recaudo de cartera",       # ingreso Wava: recaudo cuotas (sin discriminar)
    "no operativo": RUBRO_RENDIMIENTOS,      # rendimientos bancarios (rubro nuevo)
    "ajuste": "Reversas y devoluciones",     # reembolsos/reversas
}

_ACTOR_SISTEMA = "migracion:espejo-agosto-global66"


def _norm(s: object) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


def _ref_str(idt: object) -> str | None:
    if idt is None:
        return None
    if isinstance(idt, float):
        return str(int(idt))
    return str(idt).strip() or None


# ── Helpers puros (testeables sin Mongo) ───────────────────────────────────


def leer_clasificacion(path: str) -> dict[tuple[str, str, str], str]:
    """Lee la hoja clasificada → {(referencia, monto_2dec, tipo): categoria}.

    Clave compuesta (no solo referencia) para desambiguar el split Auteco: un mismo
    `ID transacción` con dos líneas de distinto valor/categoría (garantía vs préstamo).
    Excluye el footer (filas sin fecha). Falla-fuerte ante clave duplicada ambigua."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if HOJA in n), None)
    if ws is None:
        raise SystemExit(f"no se encontró la hoja '{HOJA}' en {path}")
    out: dict[tuple[str, str, str], str] = {}
    import datetime as _dt

    for row in ws.iter_rows(min_row=5, values_only=True):
        fecha = row[1] if len(row) > 1 else None
        if not isinstance(fecha, (_dt.datetime, _dt.date)):
            continue  # footer / fila sin fecha
        ref = _ref_str(row[12] if len(row) > 12 else None)
        det = row[7] if len(row) > 7 else None
        deb = row[2] if len(row) > 2 else None
        cred = row[3] if len(row) > 3 else None
        if deb not in (None, "", 0):
            monto, tipo = abs(Decimal(str(deb))), TipoMovimiento.DEBITO
        elif cred not in (None, "", 0):
            monto, tipo = abs(Decimal(str(cred))), TipoMovimiento.CREDITO
        else:
            continue
        if ref is None:
            continue
        clave = (ref, f"{monto:.2f}", tipo.value)
        if clave in out and _norm(out[clave]) != _norm(det):
            raise SystemExit(f"clasificación ambigua para {clave}: '{out[clave]}' vs '{det}'")
        out[clave] = det
    return out


def nombre_rubro_de_categoria(categoria: object) -> str:
    """Categoría (Detalle) → nombre de rubro. Alias para las 3 decisiones del CEO;
    el resto mapea a su homónimo."""
    n = _norm(categoria)
    if n in ALIAS_CATEGORIA_RUBRO:
        return ALIAS_CATEGORIA_RUBRO[n]
    return str(categoria).strip()


def verificar_totales(movimientos) -> tuple[Decimal, Decimal]:
    """Σ egresos/ingresos del parseo. Falla-fuerte si no cuadra con el footer del Excel."""
    sdeb = sum((m.monto for m in movimientos if m.tipo is TipoMovimiento.DEBITO), Decimal("0"))
    scred = sum((m.monto for m in movimientos if m.tipo is TipoMovimiento.CREDITO), Decimal("0"))
    if sdeb != SIGMA_EGRESOS or scred != SIGMA_INGRESOS:
        raise SystemExit(
            "CONTROL FAIL-LOUD: los totales del snapshot no cuadran con el Excel.\n"
            f"  egresos: {sdeb:,.2f} (esperado {SIGMA_EGRESOS:,.2f})\n"
            f"  ingresos: {scred:,.2f} (esperado {SIGMA_INGRESOS:,.2f})"
        )
    return sdeb, scred


# ── Orquestación (Mongo) ────────────────────────────────────────────────────


async def _next_orden() -> int:
    ult = await Rubro.find_all().sort(-Rubro.orden).limit(1).to_list()
    return (ult[0].orden + 1) if ult else 1


async def _asegurar_rubro_rendimientos(commit: bool) -> Rubro | None:
    r = await Rubro.find_one(Rubro.nombre == RUBRO_RENDIMIENTOS)
    if r is not None:
        print(f"[seed] rubro '{RUBRO_RENDIMIENTOS}' ya existía ({r.tipo_flujo.value}).")
        return r
    if not commit:
        print(f"[seed] (dry-run) se CREARÍA el rubro '{RUBRO_RENDIMIENTOS}' (ingreso, grupo otros).")
        return None
    nuevo = Rubro(
        grupo="otros",
        nombre=RUBRO_RENDIMIENTOS,
        tipo_flujo=TipoFlujo.INGRESO,
        orden=await _next_orden(),
        es_sistema=False,
    )
    await nuevo.insert()
    print(f"[seed] rubro '{RUBRO_RENDIMIENTOS}' creado (id {nuevo.id}).")
    return nuevo


def _construir_docs(movimientos, clasif, rubros_por_norm, mc_id):
    """MovimientoBancario[] → Transaccion[] con id_banco/ocurrencia idénticos al parser
    y rubro EXACTO del Excel. Falla-fuerte ante categoría sin rubro o tipo incoherente."""
    docs = []
    conteo: dict[tuple, int] = {}
    faltantes: set[str] = set()
    for mov in movimientos:
        clave = _clave_ocurrencia(mov)
        conteo[clave] = conteo.get(clave, 0) + 1
        cat = clasif.get((mov.referencia, f"{mov.monto:.2f}", mov.tipo.value))
        if cat is None:
            raise SystemExit(f"sin categoría en el Excel para {mov.referencia} {mov.monto} {mov.tipo.value}")
        nombre = nombre_rubro_de_categoria(cat)
        rubro = rubros_por_norm.get(_norm(nombre))
        if rubro is None:
            faltantes.add(f"{cat} → {nombre}")
            continue
        tipo_flujo = TipoFlujo.EGRESO if mov.tipo is TipoMovimiento.DEBITO else TipoFlujo.INGRESO
        if rubro["tipo_flujo"] != tipo_flujo.value:
            raise SystemExit(
                f"tipo incoherente: mov {mov.referencia} es {tipo_flujo.value} pero "
                f"rubro '{rubro['nombre']}' es {rubro['tipo_flujo']}"
            )
        docs.append(
            (
                movimiento_a_transaccion(
                    mov,
                    rubro_id=rubro["_id"],
                    mes_id=mc_id,
                    ocurrencia=conteo[clave],
                ),
                rubro["nombre"],
            )
        )
    if faltantes:
        raise SystemExit("categorías sin rubro en PROD:\n  " + "\n  ".join(sorted(faltantes)))
    return docs


async def _run(uri: str, db_name: str, path: str, commit: bool) -> None:
    if not Path(path).is_file():
        sys.exit(f"ERROR: no existe el snapshot: {Path(path).resolve()}")

    client = mongo.create_client(uri)
    await mongo.init_beanie_for(client, db_name)
    audit_service.configure_audit(client, db_name)
    modo = "COMMIT (escribe)" if commit else "DRY-RUN (solo lectura)"
    print(f"[db] conectado a '{db_name}' · modo: {modo} · snapshot: {Path(path).name}\n")

    # 1) Parseo + control fail-loud de totales.
    resultado = parse_global66(path)
    if resultado.errores:
        print(f"[parse] {len(resultado.errores)} errores de fila:")
        for e in resultado.errores[:5]:
            print(f"   fila {e.fila}: {e.motivo}")
    sdeb, scred = verificar_totales(resultado.movimientos)
    print(f"[parse] {len(resultado.movimientos)} movimientos · Σ egresos {sdeb:,.2f} · Σ ingresos {scred:,.2f} · totales OK ✅")

    # 2) Prerrequisitos.
    mc = await MesControl.find_one(MesControl.mes == MES)
    if mc is None:
        sys.exit(f"ERROR: no existe MesControl {MES[:7]} (debe estar abierto).")
    if mc.estado is EstadoMes.CERRADO:
        sys.exit(f"ERROR: {MES[:7]} está CERRADO — histórico inmutable (regla 4).")
    print(f"[mes] MesControl {MES[:7]} = {mc.estado.value} ✅")
    await _asegurar_rubro_rendimientos(commit)

    rubros_por_norm = {
        _norm(r["nombre"]): r
        for r in await Rubro.get_pymongo_collection().find(
            {}, {"nombre": 1, "tipo_flujo": 1, "es_sistema": 1, "activo": 1}
        ).to_list(length=None)
    }
    # Dry-run: el rubro nuevo aún no existe (se crea en --commit); lo inyectamos como
    # sintético para poder computar y mostrar el plan completo.
    if _norm(RUBRO_RENDIMIENTOS) not in rubros_por_norm:
        rubros_por_norm[_norm(RUBRO_RENDIMIENTOS)] = {
            "_id": PydanticObjectId(), "nombre": RUBRO_RENDIMIENTOS,
            "tipo_flujo": "ingreso", "es_sistema": False, "activo": True,
        }

    # 3) Construir docs (rubro exacto del Excel) — falla-fuerte si algo no resuelve.
    clasif = leer_clasificacion(path)
    docs = _construir_docs(resultado.movimientos, clasif, rubros_por_norm, mc.id)

    # 4) Diff contra PROD. Casos:
    #   - tx PLANA existente (id_banco): dedup; reclasificar si rubro ≠ Excel; fail-loud
    #     si el valor difiere (edición silenciosa no esperada).
    #   - referencia ya DIVIDIDA en PROD (`partes`): el Excel la trae como N líneas; si
    #     las partes de PROD == esas líneas (rubro+valor), ya está correcta → se salta
    #     TODO (ni insertar ni reclasificar). Fail-loud si no coinciden (revisar manual).
    from collections import defaultdict

    prod_por_id: dict[str, Transaccion] = {}
    refs_divididos: dict[str, set] = {}
    async for t in Transaccion.find(Transaccion.mes_id == mc.id, Transaccion.banco == Banco.GLOBAL66):
        if t.partes:
            ref = t.id_banco.split("|")[0]
            refs_divididos[ref] = {(str(p.rubro_id), f"{p.valor:.2f}") for p in t.partes}
        else:
            prod_por_id[t.id_banco] = t

    docs_por_ref: dict[str, list] = defaultdict(list)
    for d, nom in docs:
        docs_por_ref[d.id_banco.split("|")[0]].append((d, nom))

    nuevos: list = []
    a_reclasificar: list = []
    dedup = 0
    ya_divididos = 0
    for ref, grupo in docs_por_ref.items():
        if ref in refs_divididos:
            snap_set = {(str(d.rubro_id), f"{d.valor:.2f}") for d, _ in grupo}
            if snap_set == refs_divididos[ref]:
                ya_divididos += len(grupo)
                continue
            raise SystemExit(
                f"ref {ref}: las partes en PROD no coinciden con el Excel (revisar manual).\n"
                f"  PROD={sorted(refs_divididos[ref])}\n  Excel={sorted(snap_set)}"
            )
        for d, nom in grupo:
            t = prod_por_id.get(d.id_banco)
            if t is None:
                nuevos.append((d, nom))
                continue
            if t.valor != d.valor:
                raise SystemExit(
                    f"valor distinto en {d.id_banco}: PROD {t.valor:.2f} vs Excel {d.valor:.2f}"
                )
            dedup += 1
            if t.rubro_id != d.rubro_id:
                a_reclasificar.append((t, nom, d.rubro_id))
    id_a_nombre = {str(r["_id"]): r["nombre"] for r in rubros_por_norm.values()}

    n_deb = sum((d.valor for d, _ in nuevos if d.tipo_flujo is TipoFlujo.EGRESO), Decimal("0"))
    n_cred = sum((d.valor for d, _ in nuevos if d.tipo_flujo is TipoFlujo.INGRESO), Decimal("0"))

    print("\n" + "=" * 78)
    print("PLAN")
    print("=" * 78)
    print(f"  A INSERTAR (nuevos): {len(nuevos)}  ·  Σ egresos {n_deb:,.2f} · Σ ingresos {n_cred:,.2f}")
    print(f"  Ya presentes (dedup por id_banco, no se duplican): {dedup}")
    print(f"  Ya divididas en PROD y correctas (se saltan): {ya_divididos}")
    print(f"  A RECLASIFICAR (existentes con rubro ≠ Excel): {len(a_reclasificar)}")
    for tx, nom, _ in a_reclasificar:
        actual = id_a_nombre.get(str(tx.rubro_id), str(tx.rubro_id))
        print(f"      {tx.fecha}  {tx.valor:>14,.2f}  [{actual}] → [{nom}]  ({tx.descripcion[:45]})")
    print("\n  Muestra de nuevos (primeros 8):")
    for d, nom in nuevos[:8]:
        print(f"      {d.fecha}  {d.tipo_flujo.value:7s} {d.valor:>14,.2f}  [{nom}]  ({d.descripcion[:45]})")

    if not commit:
        print("\n[dry-run] no se escribió nada. Revisar y correr con --commit para aplicar.")
        client.close()
        return

    # 5) COMMIT — insertar nuevos (transacción, regla 8) + reclasificar desalineados.
    from app.transacciones.service import reclasificar_transaccion

    if nuevos:
        docs_nuevos = [d for d, _ in nuevos]
        mongo_client = Transaccion.get_pymongo_collection().database.client

        async def _finalizar(session):
            await Transaccion.insert_many(docs_nuevos, session=session)

        async with await mongo_client.start_session() as session:
            await session.with_transaction(_finalizar)
        for d in docs_nuevos:
            await emit_audit(
                AuditEvento.transaccion_creada,
                entidad="transaccion",
                entidad_id=str(d.id),
                actor_id=_ACTOR_SISTEMA,
                metadata={"origen": "migracion", "via": "espejo-agosto-global66",
                          "valor": f"{d.valor:.2f}", "tipo_flujo": d.tipo_flujo.value,
                          "id_banco": d.id_banco},
            )
        print(f"\n[commit] insertados {len(docs_nuevos)} nuevos + audit transaccion.creada.")

    for tx, nom, nuevo_rubro in a_reclasificar:
        rubro_doc = next(r for r in rubros_por_norm.values() if r["_id"] == nuevo_rubro)
        if es_rubro_clasificable(Rubro(grupo="otros", nombre=rubro_doc["nombre"], orden=0,
                                       tipo_flujo=TipoFlujo(rubro_doc["tipo_flujo"]),
                                       es_sistema=rubro_doc.get("es_sistema", False))):
            await reclasificar_transaccion(
                tx_id=str(tx.id), rubro_id=str(nuevo_rubro), usuario_id=_ACTOR_SISTEMA
            )
        else:
            # destino de sistema: set directo + audit (reclasificar lo bloquea).
            prev = tx.rubro_id
            tx.rubro_id = nuevo_rubro
            tx.clasificada_por = _ACTOR_SISTEMA
            tx.clasificada_at = now_utc()
            await tx.save()
            await emit_audit(
                AuditEvento.transaccion_clasificada, entidad="transaccion",
                entidad_id=str(tx.id), actor_id=_ACTOR_SISTEMA,
                metadata={"origen": "migracion", "rubro_anterior": str(prev),
                          "rubro_nuevo": str(nuevo_rubro)},
            )
    print(f"[commit] reclasificadas {len(a_reclasificar)} existentes.")

    # 6) Verificación post-commit.
    await _cuadre_agosto(mc.id)
    client.close()


async def _cuadre_agosto(mc_id) -> None:
    """Verificación post-commit. El Excel es SOLO Global66 → se compara la Σ de
    Global66 contra el footer; los movimientos de otros bancos / ajustes manuales
    (p. ej. la conciliación de cierre de julio) se reportan aparte (no van en el Excel)."""
    g_deb = g_cred = o_deb = o_cred = Decimal("0")
    n = ng = 0
    async for t in Transaccion.find(Transaccion.mes_id == mc_id):
        n += 1
        es_g66 = t.banco is Banco.GLOBAL66
        if es_g66:
            ng += 1
        tgt = t.valor
        if es_g66 and t.tipo_flujo is TipoFlujo.EGRESO:
            g_deb += tgt
        elif es_g66:
            g_cred += tgt
        elif t.tipo_flujo is TipoFlujo.EGRESO:
            o_deb += tgt
        else:
            o_cred += tgt
    print("\n=== CUADRE agosto en PROD (post-commit) ===")
    ok_e = "✅" if g_deb == SIGMA_EGRESOS else "❌"
    ok_i = "✅" if g_cred == SIGMA_INGRESOS else "❌"
    print(f"  Global66 ({ng} tx): Σ egresos {g_deb:,.2f} {ok_e} · Σ ingresos {g_cred:,.2f} {ok_i}")
    print(f"  vs Excel:            egresos {SIGMA_EGRESOS:,.2f}     · ingresos {SIGMA_INGRESOS:,.2f}")
    if o_deb or o_cred:
        print(f"  Otros/manuales ({n - ng} tx, fuera del Excel): egresos {o_deb:,.2f} · ingresos {o_cred:,.2f}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true", help="escribe en la BD (default: dry-run)")
    ap.add_argument("--snapshot", default=SNAPSHOT, help="ruta del snapshot clasificado")
    args = ap.parse_args()

    uri = os.environ.get("MONGODB_URI_COMPAS")
    if not uri:
        sys.exit("ERROR: falta MONGODB_URI_COMPAS (la URI NUNCA va por argv).")
    db_name = os.environ.get("MONGODB_DB", "compas")
    asyncio.run(_run(uri, db_name, args.snapshot, args.commit))


if __name__ == "__main__":
    main()
