#!/usr/bin/env python
"""Alinea los rubros de prod al Plan de Cuentas de la arquitectura presupuestal.

Lee la hoja 'Plan de Cuentas' del Excel `Compas_RODDOS_Arquitectura Presupuestal
Operativa.xlsx` (código de 3 niveles + Tipo Fijo/Variable) y POBLA en cada rubro de
prod los campos `codigo`, `grupo` (por el primer dígito del código) y `tipo`. NO
renombra: conserva los nombres cortos que ya usa el Excel de flujo (evita desincronizar
la clasificación). El grupo se corrige donde diverge (p.ej. 'Recaudo' estaba en 'otros'
→ ingresos_operativos por su código 0110).

Casos especiales (decisión CEO 2026-07-26): 'Grúas y traslados'→2130 y 'Freelance'→2140
(Operación, no estaban en el plan); 'Aportes de capital' (ingreso NO operativo, sin
código del plan) se deja en 'otros'; el 'Arriendos' DUPLICADO en grupo 'otros' se
desactiva (se conserva el de operación) si no tiene transacciones.

DRY-RUN: FLUJO_DRYRUN=1. GATE: taxonomía de datos reales → GO CEO + waiver. Idempotente.

Uso: MONGODB_URI_COMPAS=… [FLUJO_DRYRUN=1] python migrations/20260726_alinear_plan_cuentas.py \
        "docs/modelo/Compas_RODDOS_Arquitectura Presupuestal Operativa.xlsx"
"""

from __future__ import annotations

import asyncio
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, "backend")

import openpyxl  # noqa: E402

from app.db import mongo  # noqa: E402
from app.domain.rubro import Rubro, RubroGrupo, TipoRubro  # noqa: E402
from app.domain.transaccion import Transaccion  # noqa: E402

DEFAULT_XLSX = "docs/modelo/Compas_RODDOS_Arquitectura Presupuestal Operativa.xlsx"

DIG2GRUPO = {
    "0": RubroGrupo.INGRESOS_OPERATIVOS,
    "1": RubroGrupo.COSTO_PRODUCTO,
    "2": RubroGrupo.OPERACION,
    "3": RubroGrupo.NOMINA,
    "4": RubroGrupo.DEUDAS_OBLIGACIONES,
    "5": RubroGrupo.OTROS,
}
# Rubros que el CEO mantiene en Operación aunque no estén en el Plan de Cuentas.
EXTRA = {
    "gruas y traslados": ("2130", RubroGrupo.OPERACION, TipoRubro.VARIABLE),
    "freelance": ("2140", RubroGrupo.OPERACION, TipoRubro.VARIABLE),
}


def _norm(s) -> str:
    s = str(s).lower().strip()
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s+", " ", s)
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")):
        s = s.replace(a, b)
    return s


def _leer_plan(xlsx: str) -> dict[str, tuple]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Plan de Cuentas"]
    plan: dict[str, tuple] = {}
    for r in range(6, ws.max_row + 1):
        cod = ws.cell(row=r, column=2).value
        rub = ws.cell(row=r, column=3).value
        tipo = ws.cell(row=r, column=4).value
        if cod is None or rub is None:
            continue
        cod = str(cod).strip()
        if cod.endswith("000"):  # cabecera de grupo
            continue
        grupo = DIG2GRUPO[cod[0]]
        trub = TipoRubro.FIJO if str(tipo).strip().lower() == "fijo" else TipoRubro.VARIABLE
        plan[_norm(rub)] = (cod, grupo, trub)
    wb.close()
    return plan


def _resolver(nombre: str, plan: dict[str, tuple]) -> tuple | None:
    n = _norm(nombre)
    if n in EXTRA:
        return EXTRA[n]
    if n in plan:
        return plan[n]
    # prefijo: 'Producto'→'Producto (inventario…)', 'Recaudo'→'Recaudo de cartera…'
    cands = [v for pn, v in plan.items() if pn.startswith(n) or n.startswith(pn)]
    return cands[0] if len(cands) == 1 else None


async def _run(uri: str, db_name: str, xlsx: str, dry: bool) -> None:
    if not Path(xlsx).is_file():
        sys.exit(f"ERROR: no existe el Plan de Cuentas: {xlsx}")
    plan = _leer_plan(xlsx)
    client = mongo.create_client(uri)
    await mongo.init_beanie_for(client, db_name)
    print(f"[db] '{db_name}' · plan={len(plan)} rubros · {'DRY-RUN' if dry else 'APLICA'}")

    cambios = 0
    sin_match = []
    arriendos_otros = []
    async for r in Rubro.find():
        if r.nombre == "Arriendos" and r.grupo is RubroGrupo.OTROS:
            arriendos_otros.append(r)
            continue
        res = _resolver(r.nombre, plan)
        if res is None:
            sin_match.append(f"{r.nombre} ({r.grupo.value})")
            continue
        cod, grupo, trub = res
        if (r.codigo, r.grupo, r.tipo) == (cod, grupo, trub):
            continue
        print(f"  {r.nombre:42} {str(r.codigo):>5}→{cod}  {r.grupo.value}→{grupo.value}  tipo→{trub.value}")
        cambios += 1
        if not dry:
            r.codigo, r.grupo, r.tipo = cod, grupo, trub
            await r.save()

    # 'Arriendos' duplicado en 'otros': desactivar si no tiene transacciones.
    for r in arriendos_otros:
        n = await Transaccion.find(Transaccion.rubro_id == r.id).count()
        print(f"  [dup] 'Arriendos'(otros) id={r.id} txs={n} → {'desactivar' if n == 0 else 'CONSERVAR (tiene txs)'}")
        if n == 0 and not dry:
            r.activo = False
            await r.save()

    print(f"\n[resumen] cambios={cambios} · sin_match(se dejan)={sin_match}")
    if dry:
        print("[DRY-RUN] no se escribió nada.")
    client.close()


def main() -> None:
    uri = os.environ.get("MONGODB_URI_COMPAS")
    if not uri:
        sys.exit("ERROR: falta MONGODB_URI_COMPAS (nunca por argv).")
    db_name = os.environ.get("MONGODB_DB", "compas")
    dry = os.environ.get("FLUJO_DRYRUN") == "1"
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    asyncio.run(_run(uri, db_name, xlsx, dry))


if __name__ == "__main__":
    main()
