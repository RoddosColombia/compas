#!/usr/bin/env python
"""Carga curada desde el Excel 'Flujo de pagos deudas' — decisión CEO 2026-07-26.

Reemplaza la carga cruda de Global66 (revertida) por la fuente CLASIFICADA A MANO por
el CEO. Lee `Base real egresos` (con columna `Categoría`) e `Base real ingresos`, mapea
cada categoría al rubro de la arquitectura presupuestal (Plan de Cuentas) y persiste
como Transaccion. NO clasifica: la clasificación ya viene en el Excel (regla 7 estricta
en el parser). Es un tema presupuestal NETO: ingreso = neto real que entró a la cuenta;
los 193 abonos → 'Recaudo' (el desglose cuota-inicial/semanal vive en SISMO).

Decisiones CEO: 'Arriendos' → Operación; 'Grúas y traslados' y 'Freelance' → rubros
propios de Operación (ya existen en prod). id_banco = ID nativo de Global66 cuando
existe, huella determinista cuando no (idempotente). Dedup por (banco, id_banco).

DRY-RUN (no escribe): FLUJO_DRYRUN=1. GATE: datos reales → GO CEO + waiver.

Uso (URI por env var, nunca argv):
    export MONGODB_URI_COMPAS="mongodb+srv://…"
    PYTHONUTF8=1 [FLUJO_DRYRUN=1] python migrations/20260726_carga_flujo_deudas.py \
        "docs/modelo/Flujo de pagos deudas.xlsx"
"""

from __future__ import annotations

import asyncio
import os
import sys
from collections import Counter
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, "backend")

import openpyxl  # noqa: E402
from beanie import PydanticObjectId  # noqa: E402
from beanie.operators import In  # noqa: E402

from app.cargas.flujo_deudas import (  # noqa: E402
    FilaFlujoError,
    parse_fila_flujo,
    resolver_rubro_id,
)
from app.cargas.mapper import movimiento_a_transaccion  # noqa: E402
from app.db import mongo  # noqa: E402
from app.domain.bancos import Banco  # noqa: E402
from app.domain.mes_control import MesControl  # noqa: E402
from app.domain.rubro import Rubro, TipoFlujo  # noqa: E402
from app.domain.transaccion import Transaccion  # noqa: E402

DEFAULT_XLSX = "docs/modelo/Flujo de pagos deudas.xlsx"
HOJA_EGR = "Base real egresos"
HOJA_ING = "Base real ingresos"


def _filas(ws, cols: dict) -> list[dict]:
    """Lee las filas de datos de una hoja (fila 1 = headers)."""
    out = []
    for r in range(2, ws.max_row + 1):
        f = ws.cell(row=r, column=cols["fecha"]).value
        d = ws.cell(row=r, column=cols["desc"]).value
        if f is None and d is None:
            continue
        raw = {
            "fila": r,
            "fecha": f,
            "descripcion": d,
            "valor": ws.cell(row=r, column=cols["valor"]).value,
            "id_banco": ws.cell(row=r, column=cols["id"]).value,
        }
        if "cat" in cols:
            raw["categoria"] = ws.cell(row=r, column=cols["cat"]).value
        out.append(raw)
    return out


async def _mapa_rubros() -> tuple[dict[str, PydanticObjectId], PydanticObjectId]:
    """Egreso: nombre→rubro_id (en colisión, prefiere grupo 'operacion' — caso
    'Arriendos', decisión CEO). Ingreso: el único rubro de ingreso activo."""
    egreso: dict[str, PydanticObjectId] = {}
    ingreso: list[Rubro] = []
    async for r in Rubro.find(Rubro.activo == True):  # noqa: E712
        if r.tipo_flujo is TipoFlujo.EGRESO:
            if r.nombre not in egreso or r.grupo == "operacion":
                egreso[r.nombre] = r.id
        else:
            ingreso.append(r)
    if len(ingreso) != 1:
        nombres = [r.nombre for r in ingreso]
        raise SystemExit(
            f"FAIL-LOUD: se esperaba 1 rubro de ingreso activo, hay {len(ingreso)}: "
            f"{nombres}"
        )
    return egreso, ingreso[0].id


def _clave(mov) -> tuple:
    if mov.referencia:
        return (mov.banco.value, mov.referencia)
    return (mov.fecha.isoformat(), mov.tipo.value, mov.descripcion, f"{mov.monto:.2f}")


async def _run(uri: str, db_name: str, xlsx: str, dry: bool) -> None:
    ruta = Path(xlsx)
    if not ruta.is_file():
        sys.exit(f"ERROR: no existe el Excel: {ruta.resolve()}")

    client = mongo.create_client(uri)
    await mongo.init_beanie_for(client, db_name)
    from app.audit import service as audit_service

    audit_service.configure_audit(client, db_name)
    print(f"[db] '{db_name}' · Excel: {ruta.name} · {'DRY-RUN' if dry else 'CARGA REAL'}")

    egreso_map, recaudo_id = await _mapa_rubros()

    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    egr = _filas(wb[HOJA_EGR], {"fecha": 1, "desc": 2, "cat": 3, "valor": 4, "id": 6})
    ing = _filas(wb[HOJA_ING], {"fecha": 1, "desc": 2, "valor": 3, "id": 5})
    wb.close()

    # Parse + resolución de rubro + ocurrencia. Regla 7: fila mala = error, no se carga.
    errores: list[str] = []
    pend: list[tuple] = []  # (raw, mov, rubro_id, ocurrencia)
    conteo: dict[tuple, int] = {}
    for raw, tipo, rubro_fijo in (
        [(r, TipoFlujo.EGRESO, None) for r in egr]
        + [(r, TipoFlujo.INGRESO, recaudo_id) for r in ing]
    ):
        try:
            mov = parse_fila_flujo(raw, tipo_flujo=tipo)
            rubro_id = (
                rubro_fijo
                if rubro_fijo is not None
                else resolver_rubro_id(raw["categoria"], egreso_map)
            )
        except FilaFlujoError as e:
            errores.append(f"fila {raw['fila']} ({raw.get('categoria', tipo.value)}): {e}")
            continue
        k = _clave(mov)
        conteo[k] = conteo.get(k, 0) + 1
        pend.append((raw, mov, rubro_id, conteo[k]))

    # Meses presentes (derivados de la fecha, día 1) → MesControl idempotente.
    meses = sorted({f"{m.fecha.isoformat()[:7]}-01" for _, m, _, _ in pend})
    mes_id: dict[str, PydanticObjectId] = {}
    for mes in meses:
        mc = await MesControl.find_one(MesControl.mes == mes)
        if mc is None and not dry:
            mc = MesControl(mes=mes, saldo_inicial_caja=Decimal("0"))
            await mc.insert()
        # dry-run: si el mes aún no existe, id efímero (no se persiste ningún doc).
        mes_id[mes] = mc.id if mc is not None else PydanticObjectId()

    # Construir las Transaccion.
    docs: list[Transaccion] = []
    for raw, mov, rubro_id, ocurrencia in pend:
        mes = f"{mov.fecha.isoformat()[:7]}-01"
        docs.append(
            movimiento_a_transaccion(
                mov, rubro_id=rubro_id, mes_id=mes_id[mes], ocurrencia=ocurrencia
            )
        )

    # Cuadre por rubro (para el reporte).
    nombre_rubro = {r.id: r.nombre async for r in Rubro.find()}
    por_rubro: Counter = Counter()
    monto_rubro: dict = {}
    for d in docs:
        n = nombre_rubro.get(d.rubro_id, "?")
        por_rubro[n] += 1
        monto_rubro[n] = monto_rubro.get(n, Decimal("0")) + d.valor

    print(f"[parse] egresos={len(egr)} ingresos={len(ing)} · a cargar={len(docs)} · errores={len(errores)}")
    for e in errores[:10]:
        print(f"   ERROR {e}")
    print("=== cuadre por rubro (lo que entra) ===")
    for n, k in por_rubro.most_common():
        print(f"  {k:>5}  {monto_rubro[n]:>16,.0f}  {n}")

    if dry:
        print("\n[DRY-RUN] no se escribió nada en prod.")
        client.close()
        return

    # Insert idempotente: pre-filtrar (banco, id_banco) ya existentes, insertar nuevos.
    ids = [d.id_banco for d in docs]
    existentes = set()
    async for t in Transaccion.find(Transaccion.banco == Banco.GLOBAL66, In(Transaccion.id_banco, ids)):
        existentes.add(t.id_banco)
    nuevos = [d for d in docs if d.id_banco not in existentes]
    if nuevos:
        await Transaccion.insert_many(nuevos)
    print(f"\n[carga] nuevas={len(nuevos)} · duplicadas={len(docs) - len(nuevos)}")
    client.close()


def main() -> None:
    uri = os.environ.get("MONGODB_URI_COMPAS")
    if not uri:
        sys.exit("ERROR: falta la env var MONGODB_URI_COMPAS (nunca por argv).")
    db_name = os.environ.get("MONGODB_DB", "compas")
    dry = os.environ.get("FLUJO_DRYRUN") == "1"
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    asyncio.run(_run(uri, db_name, xlsx, dry))


if __name__ == "__main__":
    main()
